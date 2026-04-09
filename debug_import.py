#!/usr/bin/env python3
# Debug detalhado da importação

import os
import sys

# Deletar banco
if os.path.exists('banco.db'):
    os.remove('banco.db')
    print("✅ Banco deletado")

# Importar
from sistema import (
    get_connection, get_cursor, criar_banco, 
    normalizar_nome_coluna, get_colunas_banco, 
    preparar_colunas_extras, adicionar_coluna_dinamica
)
from openpyxl import load_workbook

print("\n=== CRIANDO BANCO ===")
conn = get_connection()

print("\n=== LENDO EXCEL ===")
wb = load_workbook('E:/planilhas-teste/celio.planilha.xlsx', read_only=True)
ws = wb.active

# Ler cabeçalhos
cabecalhos = []
for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True)):
    if cell:
        cabecalhos.append(str(cell).strip())

print(f"Cabeçalhos ({len(cabecalhos)}): {cabecalhos[:5]}...")

# Normalizar
print("\n=== NORMALIZANDO ===")
for i, cab in enumerate(cabecalhos[:5]):
    norm = normalizar_nome_coluna(cab)
    print(f"  [{i}] '{cab}' -> '{norm}'")

# Criar colunas
print("\n=== CRIANDO COLUNAS ===")
preparar_colunas_extras(cabecalhos)

# Ver colunas
print("\n=== COLUNAS NO BANCO ===")
colunas_banco = get_colunas_banco()
print(f"Total: {len(colunas_banco)}")
for c in list(colunas_banco.keys())[:10]:
    print(f"  - {c}")

# Criar mapeamento
print("\n=== MAPEAMENTO ===")
mapeamento = {}
for i, cab in enumerate(cabecalhos):
    nome_norm = normalizar_nome_coluna(cab)
    if nome_norm and nome_norm in colunas_banco:
        mapeamento[i] = nome_norm

print(f"Mapeamento: {mapeamento}")

# Ler primeira linha de dados
print("\n=== PRIMEIRA LINHA DO EXCEL ===")
for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
    print(f"Total células: {len(row)}")
    for i, cell in enumerate(row[:10]):
        val = str(cell).strip() if cell is not None else ''
        if i in mapeamento:
            print(f"  [{i}] {mapeamento[i]} = '{val}'")
        else:
            print(f"  [{i}] (não mapeado) = '{val}'")

# Testar INSERT manual
print("\n=== TESTE DE INSERT ===")
cursor = get_cursor()
colunas_insert = [c for c in colunas_banco.keys() if c != 'id']
print(f"Colunas para insert: {colunas_insert[:10]}...")

sql = f"INSERT INTO produtos ({', '.join(colunas_insert)}) VALUES ({', '.join(['?' for _ in colunas_insert])})"
print(f"SQL: {sql[:100]}...")

# Criar valores de teste
valores = {c: '' for c in colunas_insert}
valores['cliente'] = 'teste'
valores['arquivo_origem'] = 'teste.xlsx'
valores['data_importacao'] = '2024-01-01'
valores[mapeamento.get(0, 'col1')] = 'VALOR_TESTE'

tupla = tuple(valores.get(c, '') for c in colunas_insert)
print(f"Tupla: {tupla[:10]}...")

try:
    cursor.execute(sql, tupla)
    conn.commit()
    print("✅ INSERT funcionou!")
    
    # Verificar
    cursor.execute("SELECT * FROM produtos")
    dados = cursor.fetchall()
    print(f"Total no banco: {len(dados)}")
    if dados:
        print(f"Primeiro registro: {dados[0][:10]}...")
except Exception as e:
    print(f"❌ ERRO: {e}")
    import traceback
    traceback.print_exc()

wb.close()
