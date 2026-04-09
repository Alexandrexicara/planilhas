import os
import zipfile
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

def analisar_excel_completo(caminho_arquivo):
    """Análise completa do Excel para encontrar imagens"""
    print(f"=== ANÁLISE COMPLETA: {caminho_arquivo} ===")
    
    # 1. Análise da estrutura ZIP
    print("\n1. ESTRUTURA ZIP:")
    try:
        with zipfile.ZipFile(caminho_arquivo, 'r') as zip_ref:
            todos_arquivos = zip_ref.namelist()
            print(f"Total arquivos: {len(todos_arquivos)}")
            
            # Listar todas as pastas
            pastas = set()
            for f in todos_arquivos:
                if '/' in f:
                    pastas.add(f.split('/')[0])
            print(f"Pastas: {sorted(pastas)}")
            
            # Procurar imagens em todos os formatos
            extensoes = ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff', '.webp']
            imagens_todas = []
            for ext in extensoes:
                imagens_ext = [f for f in todos_arquivos if f.lower().endswith(ext)]
                imagens_todas.extend(imagens_ext)
                if imagens_ext:
                    print(f"Imagens {ext}: {len(imagens_ext)}")
                    for img in imagens_ext[:3]:  # Mostrar só 3 exemplos
                        print(f"  - {img}")
                    if len(imagens_ext) > 3:
                        print(f"  ... e mais {len(imagens_ext) - 3}")
            
            print(f"TOTAL de imagens encontradas: {len(imagens_todas)}")
            
            # Procurar por pasta media específica
            arquivos_media = [f for f in todos_arquivos if 'media' in f.lower()]
            print(f"Arquivos com 'media': {len(arquivos_media)}")
            if arquivos_media:
                print("Exemplos:")
                for f in arquivos_media[:5]:
                    print(f"  - {f}")
            
    except Exception as e:
        print(f"Erro na análise ZIP: {e}")
    
    # 2. Análise com openpyxl
    print("\n2. ANÁLISE OPENPYXL:")
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        
        print(f"Worksheet ativa: {ws.title}")
        print(f"Dimensões: {ws.max_row} x {ws.max_column}")
        
        # Verificar imagens embutidas
        if hasattr(ws, '_images'):
            print(f"Imagens embutidas: {len(ws._images)}")
            for i, img in enumerate(ws._images):
                print(f"  Imagem {i+1}: {img}")
                if hasattr(img, 'path'):
                    print(f"    Path: {img.path}")
                if hasattr(img, 'name'):
                    print(f"    Name: {img.name}")
        else:
            print("Nenhuma imagem embutida encontrada")
        
        # Verificar objetos de desenho
        if hasattr(ws, '_drawing'):
            print(f"Drawing objects: {ws._drawing}")
        
        # Verificar conteúdo das primeiras linhas
        print("\n3. CONTEÚDO DAS PRIMEIRAS LINHAS:")
        for row_num in range(1, min(6, ws.max_row + 1)):
            linha = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
            print(f"Linha {row_num}: {linha[:5]}...")  # Só 5 primeiras colunas
            
    except Exception as e:
        print(f"Erro na análise openpyxl: {e}")
    
    # 4. Verificar se há arquivos externos
    print("\n4. VERIFICAR ARQUIVOS EXTERNOS:")
    pasta_arquivo = os.path.dirname(caminho_arquivo)
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    
    # Procurar por pasta de imagens com mesmo nome
    possiveis_pastas = [
        f"{nome_base}_imagens",
        f"{nome_base}_files",
        f"{nome_base}_media",
        f"{nome_base}",
        "imagens",
        "media",
        "pictures"
    ]
    
    for pasta in possiveis_pastas:
        caminho_pasta = os.path.join(pasta_arquivo, pasta)
        if os.path.exists(caminho_pasta):
            print(f"Pasta encontrada: {caminho_pasta}")
            arquivos_pasta = os.listdir(caminho_pasta)
            imagens_pasta = [f for f in arquivos_pasta if f.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp'))]
            if imagens_pasta:
                print(f"  Imagens na pasta: {len(imagens_pasta)}")
                for img in imagens_pasta[:5]:
                    print(f"    - {img}")
            else:
                print(f"  Sem imagens na pasta")

if __name__ == "__main__":
    # Testar com os arquivos
    arquivos_teste = [
        "E:/planilhas-teste/celio.planilha.xlsx",
        "E:/planilhas-teste/05-03-2026 - 25-608_ESTUDO_250503 - COM LI ANP.xlsx"
    ]
    
    for arquivo in arquivos_teste:
        if os.path.exists(arquivo):
            analisar_excel_completo(arquivo)
            print("\n" + "="*80 + "\n")
        else:
            print(f"Arquivo não encontrado: {arquivo}")
