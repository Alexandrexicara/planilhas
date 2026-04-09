
import os
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image
from openpyxl import load_workbook

def upload_automatico(caminho_excel, cliente):
    root = tk.Tk()
    root.withdraw()  # Esconder janela principal
    
    # Verificar se já existem imagens
    pasta_cliente = os.path.join("imagens", cliente)
    if os.path.exists(pasta_cliente):
        arquivos_jpg = [f for f in os.listdir(pasta_cliente) if f.endswith('.jpg')]
        if arquivos_jpg:
            messagebox.showinfo("Info", f"Já existem {len(arquivos_jpg)} imagens para {cliente}")
            return
    
    # Perguntar se quer fazer upload
    resposta = messagebox.askyesno(
        "Upload de Imagens", 
        f"Nenhuma imagem encontrada na planilha\n\nDeseja fazer upload das imagens para o cliente '{cliente}'?"
    )
    
    if resposta:
        # Selecionar imagens
        arquivos = filedialog.askopenfilenames(
            title=f"Selecione as imagens para {cliente}",
            filetypes=[("Imagens", "*.jpg *.jpeg *.png *.gif *.bmp")]
        )
        
        if arquivos:
            # Ler produtos do Excel
            wb = load_workbook(caminho_excel, read_only=True)
            ws = wb.active
            
            codigos = []
            for row_num in range(3, min(25, ws.max_row + 1)):
                row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
                if len(row) >= 5 and row[4]:
                    codigos.append(str(row[4]).strip())
            
            wb.close()
            
            # Criar pasta se não existir
            if not os.path.exists(pasta_cliente):
                os.makedirs(pasta_cliente)
            
            # Salvar imagens com os códigos
            salvas = 0
            for i, img_path in enumerate(arquivos):
                if i < len(codigos):
                    codigo = codigos[i]
                    novo_caminho = os.path.join(pasta_cliente, f"{codigo}.jpg")
                    
                    try:
                        img = Image.open(img_path)
                        if img.mode in ('RGBA', 'LA', 'P'):
                            img = img.convert('RGB')
                        img.save(novo_caminho, 'JPEG', quality=85)
                        salvas += 1
                    except Exception as e:
                        print(f"Erro ao salvar {img_path}: {e}")
            
            messagebox.showinfo("Sucesso", f"Salvas {salvas} imagens para {cliente}!")
    
    root.destroy()

if __name__ == "__main__":
    upload_automatico("E:/planilhas-teste/celio.planilha.xlsx", "celio.planilha")
