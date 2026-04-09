import os
from openpyxl import load_workbook

def ver_linhas_excel(caminho_arquivo):
    """Mostra as primeiras linhas do Excel exatamente como estão"""
    print(f"=== ANALISANDO: {caminho_arquivo} ===")
    
    wb = load_workbook(caminho_arquivo, read_only=True)  # Sem data_only
    ws = wb.active
    
    print(f"Dimensões: {ws.max_row} x {ws.max_column}")
    print()
    
    # Mostrar primeiras 10 linhas completas
    for row_num in range(1, min(11, ws.max_row + 1)):
        linha = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
        print(f"Linha {row_num:2d}: ", end="")
        for i, cell in enumerate(linha[:10]):  # Só 10 primeiras colunas
            if cell is None:
                print(f"[None]".ljust(20), end=" ")
            else:
                valor = str(cell).strip()
                if valor == '':
                    print(f"[Vazio]".ljust(20), end=" ")
                else:
                    print(f"[{valor[:15]}...]".ljust(20), end=" ")
        print()
    
    print("\n=== ANÁLISE DA COLUNA PICTURE ===")
    for row_num in range(1, min(25, ws.max_row + 1)):
        linha = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
        if len(linha) > 0:
            picture_val = linha[0]
            print(f"Linha {row_num:2d}: PICTURE = '{picture_val}'")

if __name__ == "__main__":
    arquivo = "E:/planilhas-teste/celio.planilha.xlsx"
    if os.path.exists(arquivo):
        ver_linhas_excel(arquivo)
    else:
        print(f"Arquivo não encontrado: {arquivo}")
