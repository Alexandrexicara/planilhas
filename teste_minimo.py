#!/usr/bin/env python3
import os
import sqlite3
from openpyxl import load_workbook

# Deletar banco
if os.path.exists('banco.db'):
    os.remove('banco.db')

# Criar banco
conn = sqlite3.connect('banco.db')
cursor = conn.cursor()
cursor.execute('''
    CREATE TABLE produtos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente TEXT NOT NULL,
        arquivo_origem TEXT NOT NULL,
        data_importacao TEXT NOT NULL
    )
''')
conn.commit()

# Ler Excel
wb = load_workbook('E:/planilhas-teste/celio.planilha.xlsx', read_only=True)
ws = wb.active

# Ver primeira linha
print("=== PRIMEIRA LINHA DO EXCEL ===")
primeira = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print(f"Total de células: {len(primeira)}")
for i, cell in enumerate(primeira[:10]):
    print(f"  [{i}] = '{cell}'")

# Contar células preenchidas
preenchidas = [c for c in primeira if c]
print(f"Células preenchidas: {len(preenchidas)}")

# Se tiver poucas, procurar em outras linhas
cabecalhos = []
if len(preenchidas) <= 2:
    print("\nProcurando cabeçalhos em outras linhas...")
    for row_num in range(1, min(11, ws.max_row + 1)):
        linha = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
        celulas = [c for c in linha if c]
        print(f"  Linha {row_num}: {len(celulas)} células")
        if len(celulas) >= 5:
            cabecalhos = [str(c).strip() for c in linha if c]
            print(f"  -> Usando linha {row_num}")
            break
else:
    cabecalhos = [str(c).strip() for c in primeira if c]

print(f"\n=== CABEÇALHOS ({len(cabecalhos)}) ===")
for i, c in enumerate(cabecalhos[:10]):
    print(f"  [{i}] {c}")

# Criar colunas
print("\n=== CRIANDO COLUNAS ===")
for cab in cabecalhos:
    nome = cab.lower().replace(' ', '_').replace('-', '_')[:50]
    # Remover caracteres especiais
    import re
    nome = re.sub(r'[^\w_]', '_', nome)
    # Não começar com número
    if nome and nome[0].isdigit():
        nome = 'col_' + nome
    
    if nome and len(nome) > 1:
        try:
            cursor.execute(f'ALTER TABLE produtos ADD COLUMN "{nome}" TEXT DEFAULT ""')
            print(f"✅ {nome}")
        except Exception as e:
            print(f"❌ {nome}: {e}")

conn.commit()

# Ver colunas
print("\n=== COLUNAS NO BANCO ===")
cursor.execute("PRAGMA table_info(produtos)")
colunas = [c[1] for c in cursor.fetchall()]
print(f"Total: {len(colunas)}")
for c in colunas:
    print(f"  - {c}")

wb.close()
conn.close()
print("\n=== FIM ===")
