#!/usr/bin/env python3
# Script standalone para testar importação

import os
import sqlite3
import re
from openpyxl import load_workbook
from datetime import datetime

# Deletar banco existente
if os.path.exists('banco.db'):
    os.remove('banco.db')

# Criar banco
conn = sqlite3.connect('banco.db')
cursor = conn.cursor()
cursor.execute('''
    CREATE TABLE produtos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente TEXT NOT NULL,
        arquivo_origem TEXT NOT NULL,
        data_importacao TEXT NOT NULL
    )
''')
cursor.execute('''
    CREATE TABLE importacoes (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cliente TEXT NOT NULL,
        arquivo TEXT NOT NULL,
        data_importacao TEXT NOT NULL,
        total_registros INTEGER DEFAULT 0
    )
''')
conn.commit()

# Ler Excel
print("=== LENDO EXCEL ===")
wb = load_workbook('E:/planilhas-teste/celio.planilha.xlsx', read_only=True)
ws = wb.active

# Analisar primeiras 5 linhas
print("\n=== ANALISANDO PRIMEIRAS 5 LINHAS ===")
for row_num in range(1, 6):
    row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    cells = [c for c in row if c]
    print(f"\nLinha {row_num}: {len(cells)} células preenchidas de {len(row)} total")
    for i, cell in enumerate(row[:10]):
        if cell:
            print(f"  [{i}] = {repr(str(cell)[:40])}")

# Encontrar linha com cabeçalhos (mais de 5 células preenchidas)
print("\n=== PROCURANDO LINHA DE CABEÇALHOS ===")
cabecalhos = []
linha_cabecalho = 1
for row_num in range(1, 11):
    row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    cells = [c for c in row if c]
    if len(cells) >= 5:
        cabecalhos = [str(c).strip() for c in row if c]
        linha_cabecalho = row_num
        print(f"Cabeçalhos encontrados na linha {row_num}: {len(cabecalhos)} colunas")
        break

if not cabecalhos:
    print("ERRO: Não encontrei linha com cabeçalhos!")
    exit(1)

# Normalizar nomes
print("\n=== NORMALIZANDO NOMES ===")
def normalizar(nome):
    if not nome:
        return None
    nome = str(nome).strip().lower()
    nome = re.sub(r'[^\w\s]', '_', nome)
    nome = re.sub(r'\s+', '_', nome)
    nome = re.sub(r'_+', '_', nome)
    nome = nome.strip('_')
    if nome and nome[0].isdigit():
        nome = 'col_' + nome
    if len(nome) > 50:
        nome = nome[:50]
    if not nome:
        return None
    return nome

colunas_criadas = []
for cab in cabecalhos:
    norm = normalizar(cab)
    if norm:
        try:
            cursor.execute(f'ALTER TABLE produtos ADD COLUMN "{norm}" TEXT DEFAULT ""')
            colunas_criadas.append(norm)
            print(f"✅ {norm}")
        except Exception as e:
            print(f"❌ {norm}: {e}")

conn.commit()

# Ver colunas no banco
print(f"\n=== COLUNAS NO BANCO ({len(colunas_criadas)}) ===")

# Criar mapeamento
mapeamento = {}
for i, cab in enumerate(cabecalhos):
    norm = normalizar(cab)
    if norm in colunas_criadas:
        mapeamento[i] = norm

print(f"Mapeamento: {len(mapeamento)} colunas mapeadas")

# Inserir dados
print("\n=== INSERINDO DADOS ===")
data_atual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
cliente = 'celio.planilha'
arquivo = 'celio.planilha.xlsx'

count = 0
for row in ws.iter_rows(min_row=linha_cabecalho+1, values_only=True):
    if all(c is None or str(c).strip() == '' for c in row):
        continue
    
    valores = {'cliente': cliente, 'arquivo_origem': arquivo, 'data_importacao': data_atual}
    for i, cell in enumerate(row):
        if i in mapeamento:
            valores[mapeamento[i]] = str(cell).strip() if cell is not None else ''
    
    # Criar SQL dinâmico
    cols = list(valores.keys())
    placeholders = ', '.join(['?' for _ in cols])
    sql = f"INSERT INTO produtos ({', '.join(cols)}) VALUES ({placeholders})"
    
    try:
        cursor.execute(sql, list(valores.values()))
        count += 1
        if count == 1:
            print(f"Primeira linha: {list(valores.values())[:5]}...")
    except Exception as e:
        print(f"Erro na linha {count+1}: {e}")

conn.commit()

# Registrar importação
cursor.execute(
    "INSERT INTO importacoes (cliente, arquivo, data_importacao, total_registros) VALUES (?, ?, ?, ?)",
    (cliente, arquivo, data_atual, count)
)
conn.commit()

print(f"\n=== RESULTADO ===")
print(f"Total de produtos importados: {count}")

# Verificar
if count > 0:
    cursor.execute("SELECT * FROM produtos LIMIT 1")
    row = cursor.fetchone()
    print(f"Primeiro produto no banco: {row[:5]}...")

wb.close()
conn.close()
print("\n=== FIM ===")
