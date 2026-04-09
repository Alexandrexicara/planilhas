#!/usr/bin/env python3
"""Script de teste para debugar importação"""

import sys
sys.path.insert(0, '.')

from sistema import (
    normalizar_nome_coluna, 
    get_colunas_banco, 
    preparar_colunas_extras,
    importar_planilha,
    get_connection,
    get_cursor
)
from openpyxl import load_workbook
import os

# Deletar banco antigo
if os.path.exists('banco.db'):
    os.remove('banco.db')
    print("✅ Banco deletado")

# Forçar recriação do banco
from sistema import criar_banco, get_connection
conn = get_connection()

# Verificar estrutura inicial
print("\n=== ESTRUTURA INICIAL DO BANCO ===")
colunas = get_colunas_banco()
print(f"Colunas iniciais: {list(colunas.keys())}")

# Ler Excel e ver cabeçalhos
arquivo = 'E:/planilhas-teste/celio.planilha.xlsx'
wb = load_workbook(arquivo, read_only=True)
ws = wb.active

# Ler primeira linha
cabecalhos = []
primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
for cell in primeira_linha:
    if cell:
        cabecalhos.append(str(cell).strip())

print(f"\n=== CABEÇALHOS DO EXCEL ({len(cabecalhos)} colunas) ===")
for i, cab in enumerate(cabecalhos):
    nome_norm = normalizar_nome_coluna(cab)
    print(f"  [{i}] '{cab}' -> '{nome_norm}'")

# Preparar colunas
print("\n=== CRIANDO COLUNAS ===")
preparar_colunas_extras(cabecalhos)

# Ver colunas após criação
colunas = get_colunas_banco()
print(f"\n=== COLUNAS APÓS CRIAÇÃO ({len(colunas)}) ===")
for c in list(colunas.keys()):
    print(f"  - {c}")

# Verificar mapeamento
print("\n=== MAPEAMENTO EXCEL -> BANCO ===")
mapeamento = {}
for i, cab in enumerate(cabecalhos):
    nome_norm = normalizar_nome_coluna(cab)
    if nome_norm and nome_norm in colunas:
        mapeamento[i] = nome_norm
        print(f"  Índice {i} -> '{nome_norm}'")

# Ler dados da primeira linha de dados
print("\n=== DADOS DA PRIMEIRA LINHA DO EXCEL ===")
for row in ws.iter_rows(min_row=2, max_row=2, values_only=True):
    print(f"Total de células: {len(row)}")
    for i, cell in enumerate(row):
        valor = str(cell).strip() if cell is not None else ''
        if i in mapeamento:
            print(f"  [{i}] {mapeamento[i]}: '{valor}'")
        else:
            print(f"  [{i}] (NÃO MAPEADO): '{valor}'")

wb.close()

# Testar importação
print("\n=== TESTANDO IMPORTAÇÃO ===")
resultado = importar_planilha(arquivo)
print(f"Produtos importados: {resultado}")

# Verificar dados salvos
print("\n=== DADOS SALVOS NO BANCO ===")
cursor = get_cursor()
cursor.execute("SELECT * FROM produtos LIMIT 1")
dados = cursor.fetchone()
if dados:
    colunas_list = [c for c in colunas.keys()]
    print(f"Total de colunas: {len(colunas_list)}")
    print(f"Total de valores: {len(dados)}")
    for i, (col, val) in enumerate(zip(colunas_list, dados)):
        print(f"  {col}: '{val}'")
else:
    print("Nenhum dado encontrado!")
