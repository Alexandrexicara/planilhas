#!/usr/bin/env python3
from openpyxl import load_workbook

wb = load_workbook('E:/planilhas-teste/celio.planilha.xlsx', read_only=True)
ws = wb.active

print("=== PRIMEIRAS 10 LINHAS DO EXCEL ===")
for row_num, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
    celulas = [c for c in row if c]
    print(f"\nLinha {row_num} ({len(celulas)} células preenchidas):")
    for i, cell in enumerate(row[:15]):  # Mostrar primeiras 15 colunas
        if cell:
            print(f"  [{i}] = {repr(cell)[:50]}")

wb.close()
