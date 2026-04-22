from flask import Flask, render_template, request, jsonify, redirect, url_for, session
import sys
import os

# Adicionar o diretório atual ao path para importar o sistema
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Importar funções do sistema existente
from sistema_plus import get_connection_plus, get_cursor_plus
from extrator_imagens_excel import extrair_imagens_completas
from web_access_db import init_db, get_user_by_email
import sqlite3
import zipfile
import shutil
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'sistema_plus_integration_2024'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB

# Garantir que as pastas existam
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs('static/uploads', exist_ok=True)

# Configuração do sistema
SISTEMA_CONFIG = {
    'nome': 'Sistema Plus - Versão Web',
    'valor_original': 5000.00,
    'valor_promocional': 50.00,
    'desconto': 10,
    'versao': '2.0',
    'recursos': [
        'Importação automática de Excel',
        'Extração automática de imagens',
        'Catálogo web completo',
        'Gestão de produtos',
        'Relatórios detalhados'
    ]
}

def get_db_connection():
    """Conexão com o banco de dados PLUS existente"""
    conn = sqlite3.connect('banco_plus.db')
    conn.row_factory = sqlite3.Row
    return conn

# Funções de autenticação
def _current_user():
    """Retorna o usuário atual da sessão"""
    user_id = session.get("user_id")
    if user_id:
        return get_user_by_email(user_id)
    return None

def _is_superadmin(user):
    """Verifica se o usuário é superadmin"""
    return bool(user) and user.get("role") == "superadmin"

def _login_required(f):
    """Decorator para exigir login"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = _current_user()
        if not user:
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated_function

def _superadmin_required(f):
    """Decorator para exigir superadmin"""
    from functools import wraps
    @wraps(f)
    def decorated_function(*args, **kwargs):
        user = _current_user()
        if not _is_superadmin(user):
            return "Acesso negado - Apenas Superadmin", 403
        return f(*args, **kwargs)
    return decorated_function

def extract_images_from_excel(excel_path, output_dir):
    """Extrai imagens do arquivo Excel automaticamente"""
    images = []
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            for file in z.namelist():
                if file.startswith("xl/media/"):
                    name = os.path.basename(file)
                    path = os.path.join(output_dir, name)
                    
                    with open(path, "wb") as f:
                        f.write(z.read(file))
                    
                    images.append(path)
    except Exception as e:
        print(f"Erro ao extrair imagens: {e}")
    
    return images

def read_excel_with_images(excel_path):
    """Lê Excel e associa imagens automaticamente"""
    temp_dir = "temp_images"
    os.makedirs(temp_dir, exist_ok=True)
    
    # 1. Extrair imagens
    images = extract_images_from_excel(excel_path, temp_dir)
    
    # 2. Ler dados do Excel
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    
    rows = []
    headers = []
    
    # Pegar cabeçalhos
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # Ler linhas de dados
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = str(value) if value is not None else ""
        
        # 3. Associar imagem por ordem
        img_index = row_idx - 2  # Ajuste para índice zero
        if img_index < len(images):
            row_data['picture'] = images[img_index]
        else:
            row_data['picture'] = None
        
        rows.append(row_data)
    
    return rows, images

def save_image_to_static(image_path):
    """Move imagem para pasta static e retorna URL"""
    if not image_path or not os.path.exists(image_path):
        return None
    
    filename = os.path.basename(image_path)
    new_path = f"static/uploads/{filename}"
    
    shutil.copy(image_path, new_path)
    return f"/static/uploads/{filename}"

def import_products_to_db(products_data, filename):
    """Importa produtos para o banco de dados PLUS existente"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    imported_count = 0
    
    for product in products_data:
        try:
            # Salvar imagem se existir
            picture_url = None
            if product.get('picture'):
                picture_url = save_image_to_static(product['picture'])
            
            # Inserir no banco PLUS usando a estrutura existente
            cursor.execute("""
                INSERT INTO produtos_plus (
                    cliente, arquivo_origem, codigo, descricao, peso, 
                    valor, ncm, marca, factory, data_importacao, obs
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                product.get('cliente', 'Web'),
                filename,
                product.get('codigo', ''),
                product.get('descricao', ''),
                product.get('peso', ''),
                product.get('valor', ''),
                product.get('ncm', ''),
                product.get('marca', ''),
                product.get('factory', ''),
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                f"Importado via web: {len(products_data)} produtos"
            ))
            
            imported_count += 1
            
        except Exception as e:
            print(f"Erro ao importar produto: {e}")
            continue
    
    conn.commit()
    conn.close()
    
    # Limpar pasta temporária
    if os.path.exists("temp_images"):
        shutil.rmtree("temp_images")
    
    return imported_count

@app.route('/login')
def login_page():
    """Página de login para o sistema integration"""
    return render_template('login_integration.html', config=SISTEMA_CONFIG)

@app.route('/auth/login', methods=['POST'])
def auth_login():
    """Autenticação de login"""
    email = request.form.get("email", "").strip()
    senha = request.form.get("senha", "")
    
    user = get_user_by_email(email)
    if user and user.get("senha") == senha:  # Em produção, usar hash
        session["user_id"] = email
        return jsonify({"success": True, "redirect": url_for("index")})
    else:
        return jsonify({"success": False, "message": "Email ou senha incorretos"})

@app.route('/logout')
def logout():
    """Logout do sistema"""
    session.pop("user_id", None)
    return redirect(url_for("login_page"))

@app.route('/')
def index():
    """Página inicial atraente"""
    return render_template('index.html', config=SISTEMA_CONFIG)

@app.route('/upload')
def upload_page():
    """Página de upload"""
    return render_template('upload.html', config=SISTEMA_CONFIG)

@app.route('/api/upload', methods=['POST'])
def upload_excel():
    """API para upload e processamento de Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo inválido'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400
        
        # Salvar arquivo temporário
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(temp_path)
        
        # Processar Excel
        products_data, images = read_excel_with_images(temp_path)
        
        # Importar para banco PLUS existente
        imported_count = import_products_to_db(products_data, filename)
        
        # Limpar arquivo temporário
        os.remove(temp_path)
        
        return jsonify({
            'success': True,
            'message': f'Importação concluída com sucesso!',
            'imported_count': imported_count,
            'images_found': len(images),
            'products_count': len(products_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/catalog')
def catalog():
    """Catálogo de produtos do banco PLUS - Acesso Livre"""
    conn = get_db_connection()
    products = conn.execute("""
        SELECT * FROM produtos_plus 
        ORDER BY data_importacao DESC 
        LIMIT 100
    """).fetchall()
    conn.close()
    
    return render_template('catalog.html', products=products, config=SISTEMA_CONFIG)

@app.route('/api/stats')
def get_stats():
    """API de estatísticas do banco PLUS"""
    conn = get_db_connection()
    
    stats = {
        'total_products': conn.execute("SELECT COUNT(*) FROM produtos_plus").fetchone()[0],
        'total_clients': conn.execute("SELECT COUNT(DISTINCT cliente) FROM produtos_plus").fetchone()[0],
        'recent_imports': conn.execute("""
            SELECT COUNT(*) FROM produtos_plus 
            WHERE data_importacao >= date('now', '-7 days')
        """).fetchone()[0]
    }
    
    conn.close()
    return jsonify(stats)

@app.route('/sistema')
def sistema_info():
    """Informações sobre o sistema Python existente"""
    return render_template('sistema_info.html', config=SISTEMA_CONFIG)

if __name__ == '__main__':
    print("🚀 Iniciando Flask com integração ao Sistema Plus existente")
    print("📊 Conectado ao banco: banco_plus.db")
    print("🌐 Servidor: http://localhost:5000")
    app.run(debug=True, host='0.0.0.0', port=5000)
