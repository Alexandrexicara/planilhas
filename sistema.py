import sqlite3
import os
import json
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText
import threading
from datetime import datetime
import csv
from PIL import Image, ImageTk
import ctypes

# Base do projeto/arquivo para evitar variacao por CWD de .bat/.exe
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

from planilhas_paths import runtime_dir, ensure_from_resource, is_frozen, log_desktop

# Diretorio gravavel para DB/config/exportacoes no .exe (AppData).
DATA_DIR = runtime_dir()
DB_PATH = ensure_from_resource("banco.db") if is_frozen() else os.path.join(BASE_DIR, "banco.db")
IMAGENS_DIR = os.path.join(DATA_DIR, "imagens")

# ==============================
# BANCO DE DADOS
# ==============================

# Variáveis globais para conexão thread-safe
conn_local = threading.local()
cursor_local = threading.local()

def get_connection():
    """Obtém conexão SQLite thread-safe"""
    if not hasattr(conn_local, 'conn'):
        conn_local.conn = sqlite3.connect(DB_PATH, check_same_thread=False)
        conn_local.conn.execute("PRAGMA journal_mode=DELETE")
        conn_local.conn.execute("PRAGMA synchronous=FULL")
        conn_local.conn.execute("PRAGMA cache_size=5000")
        conn_local.conn.execute("PRAGMA temp_store=FILE")
        
        criar_banco()
    
    return conn_local.conn

def criar_banco():
    """Cria o banco de dados SQLite com 36 colunas exatas do Excel + 64 colunas vazias"""
    cursor = get_cursor()
    
    # Colunas exatas do Excel (39 colunas)
    colunas_excel = [
        "PICTURE",
        "IMAGEM",
        "LINK",
        "DOC",
        "REV",
        "ITEM",
        "CODE",
        "QUANTITY",
        "UM",
        "CCY",
        "UNIT PRICE UMO",
        "TOTAL AMOUNT UMO",
        "DESCRIÇÃO PORTUGUES (DESCRIPTION PORTUGUESE)",
        "MARCA (BRAND)",
        "INNER QUANTITY",
        "MASTER QUANTITY", 
        "TOTAL CTNS",
        "TOTAL NET WEIGHT( kg )",
        "TOTAL GROSS WEIGHT( kg )",
        "NET WEIGHT / PC( g )",
        "GROSS WEIGHT / PC( g )",
        "NET WEIGHT / CTN( kg )",
        "GROSS WEIGHT / CTN( kg )",
        "NAME OF FACTORY",
        "ADDRESS OF FACTORY",
        "TELEPHONE",
        "EAN13",
        "DUN-14 INNER",
        "DUN-14 MASTER",
        "LENGTH CTN",
        "WIDTH CTN",
        "HEIGHT CTN",
        "TOTAL CBM",
        "HS CODE",
        "PRC/KG",
        "LI",
        "OBS",
        "STATUS DA COMPRA"
    ]
    
    # Apenas colunas do Excel (sem colunas vazias adicionais)
    todas_colunas = colunas_excel
    
    # Criar SQL com aspas para nomes com espaços e caracteres especiais
    colunas_sql = ', '.join([f'"{col}" TEXT DEFAULT ""' for col in todas_colunas])
    
    cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS produtos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente TEXT NOT NULL,
            arquivo_origem TEXT NOT NULL,
            data_importacao TEXT NOT NULL,
            {colunas_sql}
        )
    """)
    
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS importacoes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            cliente TEXT NOT NULL,
            arquivo TEXT NOT NULL,
            data_importacao TEXT NOT NULL,
            total_registros INTEGER NOT NULL
        )
    """)
    
    # MIGRAÇÃO AUTOMÁTICA: Adicionar colunas IMAGEM e LINK se não existirem
    cursor.execute("PRAGMA table_info(produtos)")
    colunas_existentes = {row[1] for row in cursor.fetchall()}
    
    if 'IMAGEM' not in colunas_existentes:
        cursor.execute('ALTER TABLE produtos ADD COLUMN "IMAGEM" TEXT DEFAULT ""')
        print("✅ Migração: Coluna 'IMAGEM' adicionada ao banco de dados")
    
    if 'LINK' not in colunas_existentes:
        cursor.execute('ALTER TABLE produtos ADD COLUMN "LINK" TEXT DEFAULT ""')
        print("✅ Migração: Coluna 'LINK' adicionada ao banco de dados")
    
    get_connection().commit()
    print(f"✅ Banco criado/verificado com {len(todas_colunas)} colunas do Excel")

def get_cursor():
    """Obtém cursor thread-safe"""
    if not hasattr(cursor_local, 'cursor'):
        cursor_local.cursor = get_connection().cursor()
    return cursor_local.cursor

# ==============================
# FUNÇÕES DE IMPORTAÇÃO
# ==============================


def extrair_imagens_excel(caminho_arquivo, cliente):
    """Extrai imagens de um arquivo Excel e salva com nomes da coluna PICTURE"""
    try:
        from openpyxl import load_workbook
        from PIL import Image as PILImage
        import zipfile
        import io
        
        wb = load_workbook(caminho_arquivo, read_only=True)  # Sem data_only para ler nomes
        ws = wb.active
        
        # Criar pasta para o cliente
        pasta_cliente = os.path.join(IMAGENS_DIR, cliente)
        if not os.path.exists(pasta_cliente):
            os.makedirs(pasta_cliente)
        
        imagens_salvas = []
        
        # Método 1: Tentar extrair do ZIP do Excel
        try:
            print(f"DEBUG: Procurando imagens no ZIP do Excel...")
            with zipfile.ZipFile(caminho_arquivo, 'r') as zip_ref:
                # Procurar por arquivos de imagem
                arquivos_imagem = [f for f in zip_ref.namelist() if f.startswith('xl/media/') and f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif'))]
                print(f"DEBUG: Arquivos de imagem encontrados: {len(arquivos_imagem)}")
                print(f"DEBUG: Nomes dos arquivos: {arquivos_imagem}")
                
                # Ler primeira coluna (PICTURE) para obter nomes da linha 1
                nomes_imagens = []
                # Primeiro: pegar nome da linha 1
                linha_1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if len(linha_1) > 0 and linha_1[0]:
                    nome_img = str(linha_1[0]).strip()
                    if nome_img and nome_img != 'None':
                        nomes_imagens.append(nome_img)
                
                # Depois: pegar outros nomes das linhas de dados
                for row_num in range(3, min(25, ws.max_row + 1)):
                    row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
                    if len(row) > 0 and row[0]:
                        nome_img = str(row[0]).strip()
                        if nome_img and nome_img != 'None':
                            nomes_imagens.append(nome_img)
                
                print(f"DEBUG: Nomes da coluna PICTURE: {nomes_imagens[:5]}...")
                
                # Salvar imagens com nomes da coluna PICTURE
                for i, file in enumerate(arquivos_imagem):
                    if i < len(nomes_imagens):
                        # Extrair dados da imagem
                        image_data = zip_ref.read(file)
                        
                        # Abrir com PIL para converter se necessário
                        img = PILImage.open(io.BytesIO(image_data))
                        
                        # Usar nome da coluna PICTURE
                        nome_base = nomes_imagens[i].replace('/', '_').replace('\\', '_').replace(' ', '_')
                        nome_arquivo = f"{nome_base}.jpg"
                        caminho_salvo = os.path.join(pasta_cliente, nome_arquivo)
                        
                        # Converter para RGB se necessário
                        if img.mode in ('RGBA', 'LA', 'P'):
                            img = img.convert('RGB')
                        
                        img.save(caminho_salvo, 'JPEG', quality=85)
                        imagens_salvas.append(nome_arquivo)
                        print(f"DEBUG: Imagem salva: {caminho_salvo}")
        
        except Exception as e:
            print(f"DEBUG: Erro ao extrair imagens do ZIP: {e}")
        
        # Método 2: Tentar extrair usando openpyxl
        try:
            if hasattr(ws, '_images'):
                # Ler nomes da coluna PICTURE
                nomes_imagens = []
                # Primeiro: pegar nome da linha 1
                linha_1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
                if len(linha_1) > 0 and linha_1[0]:
                    nome_img = str(linha_1[0]).strip()
                    if nome_img and nome_img != 'None':
                        nomes_imagens.append(nome_img)
                
                # Depois: pegar outros nomes das linhas de dados
                for row_num in range(3, min(25, ws.max_row + 1)):
                    row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
                    if len(row) > 0 and row[0]:
                        nome_img = str(row[0]).strip()
                        if nome_img and nome_img != 'None':
                            nomes_imagens.append(nome_img)
                
                for i, img in enumerate(ws._images):
                    if i < len(nomes_imagens):
                        nome_base = nomes_imagens[i].replace('/', '_').replace('\\', '_').replace(' ', '_')
                        nome_arquivo = f"{nome_base}.jpg"
                        caminho_salvo = os.path.join(pasta_cliente, nome_arquivo)
                        
                        # Salvar imagem
                        img.save(caminho_salvo)
                        imagens_salvas.append(nome_arquivo)
                        print(f"DEBUG: Imagem salva via openpyxl: {caminho_salvo}")
        except Exception as e:
            print(f"DEBUG: Erro ao extrair imagens com openpyxl: {e}")
        
        print(f"DEBUG: Total de imagens extraídas: {len(imagens_salvas)}")
        print(f"DEBUG: Imagens salvas: {imagens_salvas}")
        return imagens_salvas
        
    except Exception as e:
        print(f"DEBUG: Erro geral ao extrair imagens: {e}")
        return []

def buscar_imagens_externas(caminho_arquivo, cliente):
    """Busca automaticamente imagens em pasta externa com mesmo nome do Excel"""
    try:
        from PIL import Image as PILImage
        
        # Obter pasta do arquivo Excel
        pasta_excel = os.path.dirname(caminho_arquivo)
        nome_base = os.path.basename(caminho_arquivo).replace('.xlsx', '').replace('.xls', '')
        
        # Possíveis nomes de pastas de imagens
        pastas_buscar = [
            nome_base,  # Mesmo nome do Excel
            f"{nome_base}_imagens",  # Nome + _imagens
            f"{nome_base}_images",  # Nome + _images
            f"{nome_base}_fotos",  # Nome + _fotos
            "imagens",  # Pasta genérica imagens
            "images",  # Pasta genérica images
            "fotos",  # Pasta genérica fotos
        ]
        
        imagens_encontradas = []
        
        for pasta_buscar in pastas_buscar:
            pasta_completa = os.path.join(pasta_excel, pasta_buscar)
            
            if os.path.exists(pasta_completa):
                print(f"DEBUG: Pasta encontrada: {pasta_completa}")
                
                # Listar arquivos de imagem
                arquivos_imagem = []
                for ext in ['*.jpg', '*.jpeg', '*.png', '*.gif', '*.bmp', '*.tiff']:
                    arquivos_imagem.extend(os.path.join(pasta_completa, f) for f in os.listdir(pasta_completa) if f.lower().endswith(ext.replace('*.', '.')))
                
                if arquivos_imagem:
                    # Criar pasta do cliente se não existir
                    pasta_cliente = os.path.join(IMAGENS_DIR, cliente)
                    if not os.path.exists(pasta_cliente):
                        os.makedirs(pasta_cliente)
                    
                    # Ler códigos do Excel
                    from openpyxl import load_workbook
                    wb = load_workbook(caminho_arquivo, read_only=True)
                    ws = wb.active
                    
                    codigos = []
                    for row_num in range(3, min(25, ws.max_row + 1)):
                        row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
                        if len(row) >= 5 and row[4]:
                            codigos.append(str(row[4]).strip())
                    
                    wb.close()
                    
                    # Associar imagens aos códigos (por ordem)
                    for i, img_path in enumerate(arquivos_imagem[:len(codigos)]):
                        if i < len(codigos):
                            codigo = codigos[i]
                            
                            # Converter para JPG se necessário
                            try:
                                img = PILImage.open(img_path)
                                if img.mode in ('RGBA', 'LA', 'P'):
                                    img = img.convert('RGB')
                                
                                # Salvar com nome do código
                                novo_caminho = os.path.join(pasta_cliente, f"{codigo}.jpg")
                                img.save(novo_caminho, 'JPEG', quality=85)
                                
                                imagens_encontradas.append(f"{codigo}.jpg")
                                print(f"DEBUG: Imagem salva: {novo_caminho}")
                                
                            except Exception as e:
                                print(f"DEBUG: Erro ao processar imagem {img_path}: {e}")
                    
                    return imagens_encontradas
        
        print(f"DEBUG: Nenhuma pasta de imagens encontrada para: {pastas_buscar}")
        return []

    except Exception as e:
        print(f"DEBUG: Erro ao buscar imagens externas: {e}")
        return []

def formatar_valor_celula(cell, nome_coluna=''):
    """Formata o valor da célula preservando números e preços corretamente"""
    if cell is None:
        return ''
    
    # Converter para string para verificar se é fórmula
    cell_str = str(cell).strip() if cell else ''
    
    # Se for fórmula do Excel (começa com =), tentar extrair valor numérico
    if cell_str.startswith('='):
        # Tentar calcular fórmulas simples (ex: =40*6 ou =A1*B1)
        try:
            # Remover o = e avaliar expressão matemática simples
            formula = cell_str[1:]  # Remove o =
            # Substituir referências de células por valores (aproximação)
            # Para fórmulas simples como 40*6 ou 100+50
            import re
            # Remover letras (referências de células) e manter números e operadores
            numeros = re.findall(r'\d+\.?\d*', formula)
            if len(numeros) >= 2:
                # Tentar multiplicação simples (caso mais comum: =F40*40)
                resultado = 1
                for n in numeros:
                    resultado *= float(n)
                # Verificar se é coluna de valor/preco
                coluna_lower = nome_coluna.lower() if nome_coluna else ''
                if any(palavra in coluna_lower for palavra in ['price', 'amount', 'valor', 'preco', 'preço', 'total', 'unit']):
                    return f"{resultado:.2f}"
                return str(int(resultado)) if resultado == int(resultado) else str(resultado)
        except:
            pass
        # Se não conseguir calcular, retornar vazio para fórmulas
        return ''
    
    # Se for número (int ou float)
    if isinstance(cell, (int, float)):
        # Se for coluna de preço/valor/amount, formatar com 2 casas decimais
        coluna_lower = nome_coluna.lower() if nome_coluna else ''
        # Verificar se é coluna de valor (mas não CTNS/QTY que são quantidades)
        is_valor = any(palavra in coluna_lower for palavra in ['price', 'amount', 'valor', 'preco', 'preço', 'unit'])
        is_quantidade = any(palavra in coluna_lower for palavra in ['ctns', 'qty', 'quantity', 'cartons', 'caixas'])
        
        if is_valor and not is_quantidade:
            # Formatar com 2 casas decimais, usando ponto como separador
            return f"{cell:.2f}"
        else:
            # Para outros números, converter para string sem perder precisão
            if isinstance(cell, float):
                # Remover .0 se for inteiro
                if cell == int(cell):
                    return str(int(cell))
                return str(cell)
            return str(cell)
    
    # Se for string, apenas limpar
    return cell_str

def importar_planilha(caminho_arquivo, cliente=None, progress_callback=None):
    """Importa uma única planilha (versão simplificada e funcional)"""
    if not cliente:
        cliente = os.path.basename(caminho_arquivo).replace('.xlsx', '').replace('.xls', '')
    
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(caminho_arquivo, read_only=True)
        ws = wb.active
        
        # Detectar colunas automaticamente - procurar linha com mais células preenchidas
        cabecalhos = []
        linha_cabecalho = 1
        max_celulas = 0
        
        # Procurar nas primeiras 20 linhas pela linha com mais colunas (provavelmente cabeçalhos)
        for row_num in range(2, min(21, ws.max_row + 1)):  # Começar da linha 2
            linha = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
            celulas_preenchidas = [c for c in linha if c and str(c).strip()]
            
            # Se esta linha tem mais células preenchidas que a anterior, use-a
            if len(celulas_preenchidas) > max_celulas:
                max_celulas = len(celulas_preenchidas)
                cabecalhos = [str(c).strip() for c in linha if c and str(c).strip()]
                linha_cabecalho = row_num
                
        print(f"DEBUG: Cabeçalhos encontrados na linha {linha_cabecalho}: {len(cabecalhos)} colunas")
        print(f"DEBUG: Primeiros cabeçalhos: {cabecalhos[:10]}")
        
        # DEBUG: Colunas já criadas no criar_banco(), não precisa criar dinamicamente
        # preparar_colunas_extras(cabecalhos)
        
        # Obter colunas do banco APÓS criar as dinâmicas
        colunas_banco = get_colunas_banco()
        print(f"DEBUG: Colunas no banco: {list(colunas_banco.keys())}")
        
        # Criar mapeamento: índice do Excel -> nome da coluna no banco
        # Usar nomes exatos das colunas do banco (maiúsculos)
        mapeamento = {}
        colunas_banco_list = list(colunas_banco.keys())
        colunas_nao_mapeadas = []
        
        for i, cab in enumerate(cabecalhos):
            # Procurar coluna no banco com nome exato (case-sensitive)
            if cab in colunas_banco_list:
                mapeamento[i] = cab
                print(f"✅ Coluna {i}: '{cab}' -> MAPEADA")
            elif str(cab).strip().lower() in [
                'picture', 'imagem', 'image', 'foto', 'url',
                'image url', 'url da imagem', 'link imagem', 'link da imagem'
            ]:
                mapeamento[i] = 'PICTURE'
                print(f"✅ Coluna {i}: '{cab}' -> MAPEADA (sinônimo: 'PICTURE')")
            elif str(cab).strip().lower() in [
                'imagem', 'image', 'foto', 'picture file', 'arquivo imagem'
            ]:
                mapeamento[i] = 'IMAGEM'
                print(f"✅ Coluna {i}: '{cab}' -> MAPEADA (sinônimo: 'IMAGEM')")
            elif str(cab).strip().lower() in [
                'link', 'url link', 'link url', 'imagem link', 'foto link'
            ]:
                mapeamento[i] = 'LINK'
                print(f"✅ Coluna {i}: '{cab}' -> MAPEADA (sinônimo: 'LINK')")
            else:
                # Tentar encontrar com normalização (fallback)
                nome_norm = normalizar_nome_coluna(cab)
                if nome_norm and nome_norm in colunas_banco:
                    mapeamento[i] = nome_norm
                    print(f"✅ Coluna {i}: '{cab}' -> MAPEADA (normalizado: '{nome_norm}')")
                else:
                    colunas_nao_mapeadas.append((i, cab))
                    print(f"❌ Coluna {i}: '{cab}' -> NÃO MAPEADA (não encontrada no banco)")
        
        print(f"\n=== RESUMO DO MAPEAMENTO ===")
        print(f"✅ Total mapeadas: {len(mapeamento)}")
        print(f"❌ Total não mapeadas: {len(colunas_nao_mapeadas)}")
        if colunas_nao_mapeadas:
            print(f"❌ Colunas não mapeadas: {colunas_nao_mapeadas}")
        print(f"============================\n")
        
        # Extrair imagens do Excel
        print(f"DEBUG: Extraindo imagens do arquivo: {caminho_arquivo}")
        imagens_salvas = extrair_imagens_excel(caminho_arquivo, cliente)
        print(f"DEBUG: Imagens extraídas: {len(imagens_salvas)}")
        if imagens_salvas:
            print(f"DEBUG: Nomes das imagens: {imagens_salvas[:5]}...")
        else:
            print(f"DEBUG: Nenhuma imagem encontrada no arquivo")
            
            # Buscar automática de imagens em pasta externa
            print(f"DEBUG: Buscando imagens em pasta externa...")
            imagens_externas = buscar_imagens_externas(caminho_arquivo, cliente)
            if imagens_externas:
                print(f"DEBUG: Encontradas {len(imagens_externas)} imagens externas")
                imagens_salvas = imagens_externas
            else:
                print(f"DEBUG: Nenhuma imagem externa encontrada")
        print(f"DEBUG: Total colunas banco: {len(colunas_banco)}")
        print(f"DEBUG: Colunas banco: {list(colunas_banco.keys())[:10]}")
        
        # Dados para insert
        colunas_insert = [c for c in colunas_banco.keys() if c != 'id']
        sql_colunas = ', '.join([f'"{c}"' for c in colunas_insert])
        sql_valores = ', '.join(['?' for _ in colunas_insert])
        sql_insert = f"INSERT INTO produtos ({sql_colunas}) VALUES ({sql_valores})"
        
        data_atual = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        total_importados = 0
        dados_batch = []
        batch_size = 500
        
        conn_thread = get_connection()
        cursor_thread = get_cursor()
        
        # Processar dados - ler da linha 3 em diante (SEM linha 1 como no sistema_plus.py)
        for row in ws.iter_rows(min_row=linha_cabecalho+1, values_only=True):
            # Pular linhas vazias
            if all(cell is None or str(cell).strip() == '' for cell in row):
                continue
            
            # Extrair dados usando mapeamento direto (igual ao sistema_plus.py)
            valores = {col: '' for col in colunas_banco}  # Inicializar todas colunas vazias
            
            for i, cell in enumerate(row):
                if i in mapeamento:
                    col_name = mapeamento[i]
                    # Se for PICTURE, usar exatamente da planilha
                    if col_name == 'PICTURE':
                        picture_valor = str(cell).strip() if cell is not None else ''
                        valores[col_name] = picture_valor  # ← SEM fallback, exato da planilha
                        print(f"DEBUG: PICTURE da planilha: '{picture_valor}'")
                    elif col_name == 'IMAGEM':
                        # Processar nome da imagem da coluna IMAGEM
                        imagem_valor = str(cell).strip() if cell is not None else ''
                        valores[col_name] = imagem_valor
                        print(f"DEBUG: IMAGEM da planilha: '{imagem_valor}'")
                    elif col_name == 'LINK':
                        # Processar URL de imagem da coluna LINK
                        link_valor = str(cell).strip() if cell is not None else ''
                        valores[col_name] = link_valor
                        print(f"DEBUG: LINK da planilha: '{link_valor}'")
                    else:
                        valores[col_name] = formatar_valor_celula(cell, col_name)
            
            # Calcular TOTAL AMOUNT UMO = QUANTITY × UNIT PRICE UMO
            if not valores.get('TOTAL AMOUNT UMO') and valores.get('QUANTITY') and valores.get('UNIT PRICE UMO'):
                try:
                    # Converter valores do formato brasileiro (1.200,00 → 1200.00)
                    def parse_br(valor):
                        if not valor:
                            return 0.0
                        valor_str = str(valor).strip()
                        # Remove separador de milhar (ponto) e troca vírgula por ponto decimal
                        valor_str = valor_str.replace('.', '').replace(',', '.')
                        return float(valor_str)
                    
                    quantity = parse_br(valores['QUANTITY'])
                    unit_price = parse_br(valores['UNIT PRICE UMO'])
                    total_amount = quantity * unit_price
                    
                    # Formatar resultado no padrão brasileiro (ex: 228,00)
                    if total_amount == int(total_amount):
                        valores['TOTAL AMOUNT UMO'] = f"{int(total_amount)},00"
                    else:
                        # Duas casas decimais, troca ponto por vírgula
                        valores['TOTAL AMOUNT UMO'] = f"{total_amount:.2f}".replace('.', ',')
                    
                    if total_importados < 3:
                        print(f"DEBUG: Calculo: {valores['QUANTITY']} × {valores['UNIT PRICE UMO']} = {valores['TOTAL AMOUNT UMO']}")
                except Exception as e:
                    if total_importados < 3:
                        print(f"DEBUG: Erro no cálculo: {e}")
            
            # Adicionar campos fixos (igual ao sistema_plus.py)
            valores['cliente'] = cliente
            valores['arquivo_origem'] = os.path.basename(caminho_arquivo)
            valores['data_importacao'] = data_atual
            
            # Debug da primeira linha (igual ao sistema_plus.py)
            if total_importados == 0:
                print(f"DEBUG: === PRIMEIRA LINHA SISTEMA.PY ===")
                print(f"DEBUG: Valores: {dict(list(valores.items())[:10])}")
                print(f"DEBUG: PICTURE: '{valores.get('PICTURE', '')}'")
                print(f"DEBUG: IMAGEM: '{valores.get('IMAGEM', '')}'")
                print(f"DEBUG: LINK: '{valores.get('LINK', '')}'")
            
            # Criar tupla na ordem correta
            tupla = tuple(valores.get(c, '') for c in colunas_insert)
            dados_batch.append(tupla)
            
            # Debug primeira linha - MOSTRAR TUDO
            if total_importados == 0:
                print(f"DEBUG: === PRIMEIRA LINHA ===")
                print(f"DEBUG: Valores dict: {dict(list(valores.items())[:10])}")
                print(f"DEBUG: Tupla completa: {tupla}")
                print(f"DEBUG: Primeiros 10 valores: {tupla[:10]}")
            
            total_importados += 1
            
            # Batch insert
            if len(dados_batch) >= batch_size:
                try:
                    cursor_thread.executemany(sql_insert, dados_batch)
                    conn_thread.commit()
                    print(f"DEBUG: Inseridos {len(dados_batch)} registros")
                except Exception as e:
                    print(f"DEBUG: ERRO no batch insert: {e}")
                dados_batch = []
        
        # Inserir restante
        if dados_batch:
            cursor_thread.executemany(sql_insert, dados_batch)
            conn_thread.commit()
        
        wb.close()
        
        # Registrar importação
        cursor_thread.execute(
            "INSERT INTO importacoes (cliente, arquivo, data_importacao, total_registros) VALUES (?, ?, ?, ?)",
            (cliente, os.path.basename(caminho_arquivo), data_atual, total_importados)
        )
        conn_thread.commit()
        
        print(f"DEBUG: Importados {total_importados} produtos")
        return total_importados
        
    except Exception as e:
        print(f"ERRO ao importar {caminho_arquivo}: {str(e)}")
        import traceback
        traceback.print_exc()
        return 0

def detectar_colunas_excel(cabecalhos):
    """Retorna os nomes das colunas normalizados (sem mapeamento fixo)"""
    colunas_normalizadas = {}
    
    for cabecalho in cabecalhos:
        if cabecalho:
            nome_norm = normalizar_nome_coluna(cabecalho)
            if nome_norm:
                colunas_normalizadas[nome_norm] = cabecalho
    
    return colunas_normalizadas

def get_colunas_banco():
    """Obtém lista EXATA de colunas do banco"""
    cursor_temp = get_cursor()
    cursor_temp.execute("PRAGMA table_info(produtos)")
    colunas = {}
    for col in cursor_temp.fetchall():
        nome = col[1]
        tipo = col[2]
        notnull = col[3]
        default = col[4]
        colunas[nome] = {'tipo': tipo, 'notnull': notnull, 'default': default}
    return colunas

def normalizar_nome_coluna(nome):
    """Normaliza nome da coluna para ser válido no SQLite (não pode começar com número)"""
    if not nome:
        return None
    import re
    nome = str(nome).strip()
    # Substituir caracteres especiais por underscore
    nome = re.sub(r'[^\w\s]', '_', nome)
    nome = re.sub(r'\s+', '_', nome)
    nome = re.sub(r'_+', '_', nome)
    nome = nome.strip('_')
    
    # NÃO pode começar com número no SQLite - adicionar prefixo
    if nome and nome[0].isdigit():
        nome = 'col_' + nome
    
    if len(nome) > 50:
        nome = nome[:50]
    if not nome or nome.replace('_', '').isdigit():
        return None
    return nome

def adicionar_coluna_dinamica(nome_coluna, tipo='TEXT'):
    """Adiciona uma nova coluna à tabela se não existir"""
    try:
        colunas = get_colunas_banco()
        if nome_coluna in colunas:
            return True
        
        # Limite de 200 colunas
        if len(colunas) >= 200:
            print(f"⚠️ Limite de 200 colunas atingido")
            return False
        
        cursor_temp = get_cursor()
        sql = f"ALTER TABLE produtos ADD COLUMN {nome_coluna} {tipo} DEFAULT ''"
        cursor_temp.execute(sql)
        get_connection().commit()
        print(f"✅ Coluna criada: {nome_coluna}")
        return True
    except Exception as e:
        print(f"❌ Erro ao criar coluna {nome_coluna}: {e}")
        return False

def preparar_colunas_extras(cabecalhos_excel):
    """Prepara o banco para receber todas as colunas do Excel (exceto picture/imagem)"""
    print(f"\n🔧 Preparando colunas dinâmicas...")
    
    colunas_banco = get_colunas_banco()
    colunas_adicionadas = 0
    
    # Colunas que NÃO devem ser criadas (sistema já extrai imagens separadamente)
    colunas_excluir = []
    
    for cab in cabecalhos_excel:
        nome_norm = normalizar_nome_coluna(cab)
        # Não criar colunas de imagem no banco (sistema já extrai automaticamente)
        if nome_norm and nome_norm not in colunas_banco:
            if nome_norm not in colunas_excluir:
                if adicionar_coluna_dinamica(nome_norm, 'TEXT'):
                    colunas_adicionadas += 1
    
    if colunas_adicionadas > 0:
        print(f"✅ {colunas_adicionadas} colunas extras adicionadas")
    
    return colunas_adicionadas

def limpar_banco_dados():
    """Limpa todos os dados do banco"""
    cursor = get_cursor()
    cursor.execute("DELETE FROM produtos")
    cursor.execute("DELETE FROM importacoes")
    get_connection().commit()
    print("DEBUG: Banco de dados limpo")

def importar_todas_planilhas(progress_callback=None):
    pasta = "planilhas"
    
    if not os.path.exists(pasta):
        os.makedirs(pasta)
        return 0
    
    total_geral = 0
    arquivos_processados = 0
    
    for arquivo in os.listdir(pasta):
        if arquivo.endswith(('.xlsx', '.xls')):
            caminho = os.path.join(pasta, arquivo)
            importados = importar_planilha(caminho, progress_callback=progress_callback)
            total_geral += importados
            arquivos_processados += 1
            
            if progress_callback:
                progress_callback(f"Arquivo {arquivos_processados}: {importados} produtos")
    
    return total_geral

def importar_arquivos_selecionados(arquivos, progress_callback=None):
    """Importa arquivos selecionados pelo usuário"""
    total_geral = 0
    
    for caminho in arquivos:
        importados = importar_planilha(caminho, progress_callback=progress_callback)
        total_geral += importados
        
        if progress_callback:
            progress_callback(f"Arquivo: {os.path.basename(caminho)} - {importados} produtos")
    
    return total_geral

# ==============================
# FUNÇÕES DE BUSCA
# ==============================

def buscar_produtos(termo='', cliente='Todos'):
    """Busca produtos no banco de dados - versão totalmente dinâmica"""
    # Obter colunas dinâmicas do banco
    colunas_banco = get_colunas_banco()
    colunas_select = [c for c in colunas_banco.keys() if c != 'id']
    
    # Construir query dinâmica com TODAS as colunas
    colunas_banco = get_colunas_banco()
    colunas_select = [c for c in colunas_banco.keys() if c != 'id']
    sql_colunas = ', '.join([f'"{c}"' for c in colunas_select])
    query = f"SELECT {sql_colunas} FROM produtos WHERE 1=1"
    params = []
    
    if termo and termo != '*':
        # Buscar em todas as colunas de texto
        search_cols = [c for c in colunas_select if c not in ['id']]
        if search_cols:
            conditions = ' OR '.join([f'"{c}" LIKE ?' for c in search_cols])
            query += f" AND ({conditions})"
            params.extend([f'%{termo}%'] * len(search_cols))
    
    if cliente and cliente != 'Todos':
        query += " AND cliente = ?"
        params.append(cliente)
    
    query += " ORDER BY cliente, arquivo_origem"
    
    try:
        cursor = get_cursor()
        cursor.execute(query, params)
        resultados = cursor.fetchall()
    except Exception as e:
        print(f"DEBUG: Erro na busca: {e}")
        return []
    
    return resultados

def listar_clientes():
    """Lista todos os clientes únicos (thread-safe)"""
    conn_thread = get_connection()
    cursor_thread = get_cursor()
    cursor_thread.execute("SELECT DISTINCT cliente FROM produtos ORDER BY cliente")
    clientes = [row[0] for row in cursor_thread.fetchall()]
    print(f"DEBUG: Clientes encontrados: {clientes}")
    return clientes

def contar_produtos():
    """Conta total de produtos (thread-safe)"""
    conn_thread = get_connection()
    cursor_thread = get_cursor()
    cursor_thread.execute("SELECT COUNT(*) FROM produtos")
    total = cursor_thread.fetchone()[0]
    print(f"DEBUG: Total de produtos: {total}")
    return total

def contar_importacoes():
    """Conta total de importações (thread-safe)"""
    conn_thread = get_connection()
    cursor_thread = get_cursor()
    cursor_thread.execute("SELECT COUNT(*) FROM importacoes")
    return cursor_thread.fetchone()[0]

# ==============================
# FUNÇÕES DE EXPORTAÇÃO
# ==============================

def normalizar_pasta_exportacoes(pasta):
    """Normaliza pasta de exportacoes e evita exportacoes/exportacoes/..."""
    pasta_base = os.path.normpath(pasta) if pasta else 'exportacoes'
    drive, tail = os.path.splitdrive(pasta_base)
    is_abs = tail.startswith(os.sep)
    partes = [p for p in tail.split(os.sep) if p]
    
    partes_limpas = []
    for parte in partes:
        if (
            partes_limpas
            and partes_limpas[-1].lower() == 'exportacoes'
            and parte.lower() == 'exportacoes'
        ):
            continue
        partes_limpas.append(parte)
    
    if not partes_limpas:
        partes_limpas = ['exportacoes']
    elif partes_limpas[-1].lower() != 'exportacoes':
        partes_limpas.append('exportacoes')
    
    prefixo = ''
    if drive:
        prefixo = drive + (os.sep if is_abs else '')
    elif is_abs:
        prefixo = os.sep
    
    return prefixo + os.sep.join(partes_limpas)

def sanitizar_nome_arquivo(nome):
    """Gera um nome de arquivo seguro no Windows."""
    import re
    if nome is None:
        return ''
    nome = str(nome).strip()
    # Remover caracteres proibidos no Windows: <>:"/\\|?* e controles
    nome = re.sub(r'[<>:"/\\|?*\x00-\x1f]', '_', nome)
    nome = re.sub(r'\\s+', ' ', nome).strip()
    nome = nome.strip(' ._')
    if len(nome) > 120:
        nome = nome[:120].rstrip(' ._')
    return nome

def exportar_resultados(resultados, formato='excel', pasta_exportacoes=None, colunas=None):
    """Exporta resultados da busca (versão sem pandas) - Exporta apenas planilha do filtro atual"""
    if not resultados:
        print("DEBUG EXPORT PY: sem resultados, exportacao cancelada")
        return None
    
    # Usar pasta configurada ou pasta padrao
    if pasta_exportacoes is None:
        pasta_exportacoes = os.path.join(DATA_DIR, "exportacoes")
    
    print(f"DEBUG EXPORT PY: cwd={os.getcwd()}")
    print(f"DEBUG EXPORT PY: base_dir={BASE_DIR}")
    print(f"DEBUG EXPORT PY: pasta_destino={pasta_exportacoes}")
    print(f"DEBUG EXPORT PY: formato={formato} | linhas={len(resultados)}")
    if is_frozen():
        log_desktop(f"DEBUG EXPORT PY: pasta_destino={pasta_exportacoes} formato={formato} linhas={len(resultados)}")
    
    # Criar pasta exportacoes automaticamente
    os.makedirs(pasta_exportacoes, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

    header = None
    if colunas:
        header = [str(c) for c in colunas]
        max_len = max((len(r) for r in resultados), default=0)
        
        # DEBUG: Mostrar informações das colunas
        print(f"DEBUG EXPORT: header={header[:10]}... (total: {len(header)})")
        print(f"DEBUG EXPORT: max_len dos resultados={max_len}")
        print(f"DEBUG EXPORT: primeira linha={resultados[0] if resultados else 'vazio'}")
        
        if max_len and len(header) > max_len:
            print(f"DEBUG EXPORT: Header truncado de {len(header)} para {max_len}")
            header = header[:max_len]
        elif max_len and len(header) < max_len:
            header = header + [f"COL_{i}" for i in range(len(header) + 1, max_len + 1)]

    # Exportar apenas a planilha selecionada no filtro
    if formato in ('excel', 'csv') and header and 'ARQUIVO_ORIGEM' in header and 'CLIENTE' in header:
        idx_cliente = header.index('CLIENTE')
        idx_arquivo = header.index('ARQUIVO_ORIGEM')
        
        # Pegar a planilha atual do filtro (primeira linha dos resultados)
        if resultados:
            primeira_linha = resultados[0]
            cliente_atual = primeira_linha[idx_cliente] if idx_cliente < len(primeira_linha) else ''
            planilha_atual = primeira_linha[idx_arquivo] if idx_arquivo < len(primeira_linha) else ''
            
            # Exportar apenas essa planilha
            base_planilha = os.path.splitext(str(planilha_atual or 'resultado'))[0]
            base_planilha = sanitizar_nome_arquivo(base_planilha) or 'resultado'
            base_cliente = sanitizar_nome_arquivo(cliente_atual) or 'cliente'
            nome_saida = f"{base_planilha}__{base_cliente}__{timestamp}.csv"
            caminho = os.path.join(pasta_exportacoes, nome_saida)
            
            # Evitar sobrescrever
            if os.path.exists(caminho):
                n = 2
                while True:
                    alt = os.path.join(pasta_exportacoes, f"{base_planilha}__{base_cliente}__{timestamp}_{n}.csv")
                    if not os.path.exists(alt):
                        caminho = alt
                        break
                    n += 1
            
            with open(caminho, 'w', newline='', encoding='utf-8-sig') as f:
                writer = csv.writer(f, delimiter=';')
                writer.writerow(header)
                writer.writerows(resultados)
                f.flush()
                os.fsync(f.fileno())
            
            if not os.path.exists(caminho):
                raise FileNotFoundError(f"Arquivo nao foi criado: {caminho}")
            
            print(f"DEBUG EXPORT PY: arquivo={caminho} | criado=True | tamanho={os.path.getsize(caminho)}")
            if is_frozen():
                log_desktop(f"DEBUG EXPORT PY: arquivo={caminho} criado=True tamanho={os.path.getsize(caminho)}")
            
            return caminho
    
    # Fallback: exportar como CSV simples
    if formato in ('excel', 'csv'):
        arquivo = os.path.join(pasta_exportacoes, f"resultados_busca_{timestamp}.csv")
        with open(arquivo, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            if header:
                writer.writerow(header)
            else:
                writer.writerow(['Cliente', 'Arquivo', 'Código', 'Descrição', 'Peso', 'Valor', 'NCM'])
            writer.writerows(resultados)
            f.flush()
            os.fsync(f.fileno())
        print(f"DEBUG EXPORT PY: arquivo={arquivo} | criado={os.path.exists(arquivo)} | tamanho={os.path.getsize(arquivo) if os.path.exists(arquivo) else -1}")
        if is_frozen():
            log_desktop(f"DEBUG EXPORT PY: arquivo={arquivo} criado={os.path.exists(arquivo)} tamanho={os.path.getsize(arquivo) if os.path.exists(arquivo) else -1}")
        return arquivo
    
    return None

# ==============================
# INTERFACE GRÁFICA
# ==============================

class SistemaPlanilhas:
    def __init__(self):
        # Configurar ID do aplicativo ANTES de criar a janela (para ícone na taskbar)
        try:
            if os.name == 'nt':  # Windows
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID('planilhas.com.sistema')
        except Exception as e:
            print(f"DEBUG: Erro ao configurar AppUserModelID: {e}")
        
        self.janela = tk.Tk()
        self.janela.title("Sistema Profissional de Planilhas - Busca Rápida")
        self.janela.geometry("1200x700")
        self.janela.configure(bg='#f0f0f0')
        
        # Ícone da janela
        try:
            icon_path = os.path.join(BASE_DIR, "icon.ico")
            if os.path.exists(icon_path):
                self.janela.iconbitmap(icon_path)
        except Exception as e:
            print(f"DEBUG: Erro ao configurar ícone: {e}")
        
        # Variáveis
        self.termo_busca = tk.StringVar()
        self.cliente_selecionado = tk.StringVar(value="Todos")
        self.pasta_exportacoes = self.carregar_configuracao_exportacoes()
        
        self.criar_interface()
        self.atualizar_estatisticas()
        self.atualizar_lista_clientes()
    
    def carregar_configuracao_exportacoes(self):
        """Carrega a configuracao da pasta exportacoes de um arquivo JSON"""
        arquivo_config = os.path.join(DATA_DIR, "config_exportacoes.json")
        pasta_padrao = os.path.join(DATA_DIR, "exportacoes")
        
        if os.path.exists(arquivo_config):
            try:
                with open(arquivo_config, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    pasta = config.get('pasta_exportacoes', pasta_padrao)
                    pasta_norm = os.path.abspath(normalizar_pasta_exportacoes(pasta))
                    os.makedirs(pasta_norm, exist_ok=True)
                    print(f"DEBUG CONFIG PY: arquivo_config={arquivo_config}")
                    print(f"DEBUG CONFIG PY: pasta_raw={pasta}")
                    print(f"DEBUG CONFIG PY: pasta_norm={pasta_norm}")
                    
                    # Corrige automaticamente caminhos antigos quebrados
                    if pasta != pasta_norm:
                        self.salvar_configuracao_exportacoes(pasta_norm)
                    
                    return pasta_norm
            except Exception as e:
                print(f"DEBUG: Erro ao carregar configuracao: {e}")
        
        os.makedirs(pasta_padrao, exist_ok=True)
        return pasta_padrao
    
    def salvar_configuracao_exportacoes(self, pasta):
        """Salva a configuracao da pasta exportacoes em um arquivo JSON"""
        arquivo_config = os.path.join(DATA_DIR, "config_exportacoes.json")
        try:
            pasta_norm = os.path.abspath(normalizar_pasta_exportacoes(pasta))
            with open(arquivo_config, 'w', encoding='utf-8') as f:
                json.dump({'pasta_exportacoes': pasta_norm}, f, indent=4)
            print(f"DEBUG: Configuracao salva: {pasta_norm}")
            print(f"DEBUG CONFIG PY: arquivo_config={arquivo_config}")
        except Exception as e:
            print(f"DEBUG: Erro ao salvar configuracao: {e}")
    
    def escolher_pasta_exportacoes(self):
        """Abre dialogo para escolher pasta de exportacoes"""
        from tkinter import filedialog, messagebox
        
        pasta_escolhida = filedialog.askdirectory(
            title="Escolha onde criar a pasta de exportações",
            initialdir=self.pasta_exportacoes
        )
        
        if pasta_escolhida:
            pasta_completa = os.path.abspath(normalizar_pasta_exportacoes(pasta_escolhida))
            os.makedirs(pasta_completa, exist_ok=True)
            print(f"DEBUG CONFIG PY: pasta_escolhida={pasta_escolhida}")
            print(f"DEBUG CONFIG PY: pasta_final={pasta_completa}")
            
            self.pasta_exportacoes = pasta_completa
            self.salvar_configuracao_exportacoes(pasta_completa)
            
            messagebox.showinfo("Sucesso", 
                f"Pasta de exportações configurada em:\n{pasta_completa}\n\n"
                "Todas as exportações serão salvas automaticamente aqui.")
    
    def criar_interface(self):
        # Frame superior - Estatísticas (mais estreito)
        frame_stats = tk.Frame(self.janela, bg='#2c3e50', height=90)
        frame_stats.pack(fill='x', padx=5, pady=5)
        frame_stats.pack_propagate(False)
        
        # Frame para logo e título (linha superior)
        frame_logo_titulo = tk.Frame(frame_stats, bg='#2c3e50')
        frame_logo_titulo.pack(fill='x', pady=(5, 2))
        
        # Logo no canto esquerdo (menor)
        try:
            logo_path = os.path.join(BASE_DIR, "img", "Penacho laranja em fundo neutro.png")
            if os.path.exists(logo_path):
                # Carregar e redimensionar imagem (menor)
                logo_image = Image.open(logo_path)
                logo_image = logo_image.resize((50, 50), Image.Resampling.LANCZOS)
                logo_photo = ImageTk.PhotoImage(logo_image)
                
                logo_label = tk.Label(frame_logo_titulo, image=logo_photo, bg='#2c3e50')
                logo_label.image = logo_photo  # Manter referência
                logo_label.pack(side='left', padx=(15, 10))
            else:
                # Fallback para emoji se imagem não existir
                logo_label = tk.Label(frame_logo_titulo, text="🪶", font=('Arial', 36), bg='#2c3e50', fg='white')
                logo_label.pack(side='left', padx=(15, 10))
        except Exception as e:
            # Fallback para emoji em caso de erro
            logo_label = tk.Label(frame_logo_titulo, text="🪶", font=('Arial', 36), bg='#2c3e50', fg='white')
            logo_label.pack(side='left', padx=(15, 10))
        
        # Título ao lado do logo (fonte menor)
        titulo_label = tk.Label(frame_logo_titulo, text="planilhas.com", font=('Arial', 18, 'bold'), bg='#2c3e50', fg='white')
        titulo_label.pack(side='left', padx=(0, 15))
        
        # Frame de navegação no canto direito
        frame_navegacao = tk.Frame(frame_logo_titulo, bg='#2c3e50')
        frame_navegacao.pack(side='right', padx=(20, 0))
        
        # Botão para ir para o Sistema Plus
        btn_ir_para_plus = tk.Button(frame_navegacao, text="⭐ Sistema PLUS", 
                                   command=self.ir_para_sistema_plus,
                                   bg='#9b59b6', fg='white', font=('Arial', 10, 'bold'),
                                   relief='raised', bd=2, cursor='hand2')
        btn_ir_para_plus.pack(pady=2)
        
        # Frame de estatísticas (linha inferior do cabeçalho) - mais espaço para não cortar
        frame_stats_container = tk.Frame(frame_stats, bg='#2c3e50')
        frame_stats_container.pack(fill='x', pady=(5, 8))
        
        self.label_stats = tk.Label(frame_stats_container, text="", bg='#2c3e50', fg='white', 
                                   font=('Arial', 11, 'bold'), height=2)
        self.label_stats.pack(fill='x', padx=10)
        
        # Frame de busca
        frame_busca = tk.LabelFrame(self.janela, text="🔍 Busca Avançada", font=('Arial', 12, 'bold'), bg='#f0f0f0')
        frame_busca.pack(fill='x', padx=10, pady=5)
        
        # Container para alinhamento perfeito
        frame_busca_container = tk.Frame(frame_busca, bg='#f0f0f0')
        frame_busca_container.pack(fill='x', padx=10, pady=10)
        
        # Linha 1 - Campo de busca
        frame_busca_linha1 = tk.Frame(frame_busca_container, bg='#f0f0f0')
        frame_busca_linha1.pack(fill='x', pady=(0, 5))
        
        tk.Label(frame_busca_linha1, text="Buscar:", font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#2c3e50').pack(side='left', padx=(0, 10))
        self.campo_busca = tk.Entry(frame_busca_linha1, textvariable=self.termo_busca, font=('Arial', 10), width=40, relief='solid', bd=1)
        self.campo_busca.pack(side='left', padx=(0, 10), fill='x', expand=True)
        self.campo_busca.bind('<Return>', lambda e: self.executar_busca())
        
        # Botões de busca alinhados
        frame_botoes_busca = tk.Frame(frame_busca_linha1, bg='#f0f0f0')
        frame_botoes_busca.pack(side='right')
        
        tk.Button(frame_botoes_busca, text="🔎 Buscar", command=self.executar_busca, 
                 bg='#3498db', fg='white', font=('Arial', 10, 'bold'), width=12, height=1,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=(0, 5))
        
        tk.Button(frame_botoes_busca, text="🗑️ Limpar", command=self.limpar_busca,
                 bg='#e74c3c', fg='white', font=('Arial', 10, 'bold'), width=12, height=1,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=(0, 5))
        
        # Linha 2 - Filtros
        frame_busca_linha2 = tk.Frame(frame_busca_container, bg='#f0f0f0')
        frame_busca_linha2.pack(fill='x', pady=(5, 0))
        
        tk.Label(frame_busca_linha2, text="Cliente:", font=('Arial', 10, 'bold'), bg='#f0f0f0', fg='#2c3e50').pack(side='left', padx=(0, 10))
        self.combo_clientes = ttk.Combobox(frame_busca_linha2, textvariable=self.cliente_selecionado, 
                                          state='readonly', width=30, font=('Arial', 10))
        self.combo_clientes.pack(side='left')
        
        # Frame único para ambas as categorias na mesma linha
        frame_categorias = tk.Frame(self.janela, bg='#f0f0f0')
        frame_categorias.pack(fill='x', padx=10, pady=5)
        
        # Frame de importação (lado esquerdo)
        frame_import = tk.LabelFrame(frame_categorias, text="📂 Importação de Planilhas", font=('Arial', 12, 'bold'), bg='#f0f0f0')
        frame_import.pack(side='left', fill='both', expand=True, padx=(0, 5), pady=5)
        
        # Container para botões de importação
        frame_botoes_import = tk.Frame(frame_import, bg='#f0f0f0')
        frame_botoes_import.pack(fill='x', padx=10, pady=10)
        
        # Botões de importação em uma linha
        tk.Button(frame_botoes_import, text="📁 Importar Pasta", command=self.importar_pasta,
                 bg='#27ae60', fg='white', font=('Arial', 9, 'bold'), width=18, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_botoes_import, text="📄 Selecionar Arquivos", command=self.selecionar_arquivos,
                 bg='#f39c12', fg='white', font=('Arial', 9, 'bold'), width=18, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_botoes_import, text="🔄 Limpar Banco", command=self.limpar_banco,
                 bg='#e67e22', fg='white', font=('Arial', 9, 'bold'), width=18, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        # Frame de visão geral (lado direito)
        frame_visao = tk.LabelFrame(frame_categorias, text="🎯 Visão Geral", font=('Arial', 12, 'bold'), bg='#f0f0f0')
        frame_visao.pack(side='left', fill='both', expand=True, padx=(5, 0), pady=5)
        
        # Container para botões de visão geral
        frame_botoes_visao = tk.Frame(frame_visao, bg='#f0f0f0')
        frame_botoes_visao.pack(fill='x', padx=10, pady=10)
        
        # Botões de visão geral em uma linha
        tk.Button(frame_botoes_visao, text="📊 Capacidade", command=self.mostrar_capacidade,
                 bg='#9b59b6', fg='white', font=('Arial', 9, 'bold'), width=18, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_botoes_visao, text="📖 Manual", command=self.abrir_manual,
                 bg='#FF8C00', fg='white', font=('Arial', 9, 'bold'), width=18, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_botoes_visao, text="🚀 PLUS", command=self.abrir_sistema_plus,
                 bg='#e74c3c', fg='white', font=('Arial', 9, 'bold'), width=18, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        # Frame de resultados
        frame_resultados = tk.LabelFrame(self.janela, text="📊 Resultados da Busca", font=('Arial', 12, 'bold'), bg='#f0f0f0')
        frame_resultados.pack(fill='both', expand=True, padx=10, pady=5)
        
        # Frame de exportação
        frame_export = tk.Frame(frame_resultados, bg='#f0f0f0')
        frame_export.pack(fill='x', padx=5, pady=5)
        
        tk.Button(frame_export, text="📊 Exportar Excel", command=self.exportar_excel,
                 bg='#8e44ad', fg='white', font=('Arial', 10, 'bold'), width=15, height=1,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_export, text="📄 Exportar CSV", command=self.exportar_csv,
                 bg='#16a085', fg='white', font=('Arial', 10, 'bold'), width=15, height=1,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_export, text="📁 Pasta Exportações", command=self.escolher_pasta_exportacoes,
                 bg='#3498db', fg='white', font=('Arial', 10, 'bold'), width=18, height=1,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_export, text="🖼️ Ver Foto", command=self.ver_foto,
                 bg='#e67e22', fg='white', font=('Arial', 10, 'bold'), width=12, height=1,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        self.label_resultados = tk.Label(frame_export, text="0 resultados encontrados", bg='#f0f0f0', font=('Arial', 10))
        self.label_resultados.pack(side='right', padx=5)
        
        # CONTAINER DE COLUNAS - Compacto
        frame_colunas = tk.LabelFrame(frame_resultados, text="📋 Colunas Visíveis", 
                                      font=('Arial', 10, 'bold'), bg='#d5d8dc', fg='#2c3e50')
        frame_colunas.pack(fill='x', padx=5, pady=2)
        
        # Frame scrollável para os checkboxes (altura reduzida)
        canvas_colunas = tk.Canvas(frame_colunas, bg='#d5d8dc', height=50, highlightthickness=0)
        scroll_colunas = ttk.Scrollbar(frame_colunas, orient='horizontal', command=canvas_colunas.xview)
        frame_checks = tk.Frame(canvas_colunas, bg='#d5d8dc')
        
        canvas_colunas.configure(xscrollcommand=scroll_colunas.set)
        canvas_colunas.pack(fill='x', expand=True)
        scroll_colunas.pack(fill='x')
        
        canvas_colunas.create_window((0, 0), window=frame_checks, anchor='nw')
        
        # Atualizar scrollregion quando o conteúdo mudar
        def atualizar_scrollregion(event=None):
            canvas_colunas.configure(scrollregion=canvas_colunas.bbox('all'))
        
        frame_checks.bind('<Configure>', atualizar_scrollregion)
        
        # Permitir scroll com mouse
        def on_mousewheel(event):
            canvas_colunas.xview_scroll(int(-1*(event.delta/120)), 'units')
        
        canvas_colunas.bind('<MouseWheel>', on_mousewheel)
        frame_checks.bind('<MouseWheel>', on_mousewheel)
        
        # Tabela de resultados
        frame_tabela = tk.Frame(frame_resultados)
        frame_tabela.pack(fill='both', expand=True, padx=5, pady=5)
        
        # Obter colunas dinâmicas do banco para a tabela
        try:
            colunas_banco = get_colunas_banco()
            colunas_base = [c for c in colunas_banco.keys() if c != 'id']
            
            # Reordenar para IMAGEM ficar ao lado de DESCRICAO
            colunas_ordenadas = []
            descricao_idx = -1
            for i, c in enumerate(colunas_base):
                if c == 'descricao':
                    descricao_idx = i
                    break
            
            if descricao_idx >= 0:
                # Inserir colunas até descricao (inclusive)
                colunas_ordenadas = colunas_base[:descricao_idx+1]
                # Adicionar IMAGEM
                colunas_ordenadas.append('IMAGEM')
                # Adicionar restante das colunas
                colunas_ordenadas.extend(colunas_base[descricao_idx+1:])
            else:
                colunas_ordenadas = colunas_base + ['IMAGEM']
            
            self.colunas = tuple([c.upper() for c in colunas_ordenadas])
        except Exception as e:
            # Fallback para colunas padrão com IMAGEM e LINK ao lado de DESCRICAO
            self.colunas = ('CLIENTE', 'ARQUIVO_ORIGEM', 'CODIGO', 'DESCRICAO', 'IMAGEM', 'LINK', 'PESO', 'VALOR', 'NCM')
        
        self.tabela = ttk.Treeview(frame_tabela, columns=self.colunas, show='headings', height=15)
        
        # Configurar colunas dinamicamente - todas com stretch para mostrar conteudo completo
        for col in self.colunas:
            self.tabela.heading(col, text=col.replace('_', ' '))
            if 'DESCRI' in col:
                # Descricao bem larga para texto longo
                self.tabela.column(col, width=800, minwidth=300, stretch=True)
            elif col == 'CLIENTE':
                self.tabela.column(col, width=200, minwidth=150, stretch=True)
            elif col in ('ARQUIVO_ORIGEM',):
                self.tabela.column(col, width=250, minwidth=150, stretch=True)
            elif col in ('CODE', 'ITEM'):
                self.tabela.column(col, width=150, minwidth=100, stretch=True)
            elif col == 'IMAGEM':
                self.tabela.column(col, width=80, minwidth=60, stretch=False)
            elif col == 'LINK':
                self.tabela.column(col, width=200, minwidth=150, stretch=True)
            else:
                # Todas as outras colunas com largura generosa e stretch
                self.tabela.column(col, width=200, minwidth=120, stretch=True)
        
        # Scrollbars
        scroll_v = ttk.Scrollbar(frame_tabela, orient='vertical', command=self.tabela.yview)
        scroll_h = ttk.Scrollbar(frame_tabela, orient='horizontal', command=self.tabela.xview)
        self.tabela.configure(yscrollcommand=scroll_v.set, xscrollcommand=scroll_h.set)
        
        self.tabela.grid(row=0, column=0, sticky='nsew')
        scroll_v.grid(row=0, column=1, sticky='ns')
        scroll_h.grid(row=1, column=0, sticky='ew')
        
        frame_tabela.grid_rowconfigure(0, weight=1)
        frame_tabela.grid_rowconfigure(1, weight=0)
        frame_tabela.grid_columnconfigure(0, weight=1)
        frame_tabela.grid_columnconfigure(1, weight=0)
        
        # Variáveis para checkboxes das colunas
        self.colunas_visiveis = {}
        self.checks_colunas = {}
        
        # Criar checkboxes para cada coluna
        colunas_disponiveis = list(self.colunas)
        
        # Carregar configuração salva ou usar padrão
        colunas_salvas = self.carregar_configuracao_colunas()
        if colunas_salvas:
            colunas_padrao = colunas_salvas
        else:
            colunas_padrao = ['CLIENTE', 'CODIGO', 'DESCRICAO', 'IMAGEM', 'VALOR']  # Colunas padrão selecionadas
        
        for i, col in enumerate(colunas_disponiveis):
            var = tk.BooleanVar(value=col in colunas_padrao)
            self.colunas_visiveis[col] = var
            
            chk = tk.Checkbutton(frame_checks, text=col.replace('_', ' '), variable=var,
                                bg='#d5d8dc', font=('Arial', 9), fg='#2c3e50',
                                selectcolor='#ffffff',
                                command=lambda c=col: self.atualizar_colunas_visiveis())
            chk.pack(side='left', padx=5, pady=2)
            self.checks_colunas[col] = chk
        
        # Botão para aplicar seleção
        btn_aplicar = tk.Button(frame_colunas, text="✓ Aplicar", 
                               command=self.atualizar_colunas_visiveis,
                               bg='#27ae60', fg='white', font=('Arial', 8), width=8)
        btn_aplicar.pack(side='right', padx=5, pady=2)
        
        # Atualizar scrollregion
        frame_checks.update_idletasks()
        canvas_colunas.configure(scrollregion=canvas_colunas.bbox('all'))
        
        # Aplicar configuração inicial após criar a tabela
        self.janela.after(100, self.atualizar_colunas_visiveis)
        
        # Barra de status
        self.status_bar = tk.Label(self.janela, text="Pronto", bd=1, relief='sunken', anchor='w', bg='#ecf0f1')
        self.status_bar.pack(side='bottom', fill='x')
    
    def atualizar_estatisticas(self):
        """Atualiza as estatísticas na interface"""
        total_produtos = contar_produtos()
        total_importacoes = contar_importacoes()
        
        stats_text = f"📦 Produtos: {total_produtos:,}  |  📁 Importações: {total_importacoes:,}  |  🕒 Última atualização: {datetime.now().strftime('%H:%M:%S')}"
        self.label_stats.config(text=stats_text)
    
    def atualizar_colunas_visiveis(self):
        """Atualiza quais colunas são visíveis na tabela baseado nos checkboxes"""
        # Obter colunas selecionadas
        colunas_selecionadas = [col for col, var in self.colunas_visiveis.items() if var.get()]
        
        if not colunas_selecionadas:
            messagebox.showwarning("Aviso", "Selecione pelo menos uma coluna!")
            return
        
        # Ocultar/mostrar colunas na tabela
        for col in self.colunas:
            if col in colunas_selecionadas:
                largura = self._get_largura_coluna(col)
                minwidth = self._get_minwidth_coluna(col)
                self.tabela.column(col, width=largura, minwidth=minwidth, stretch=True)
            else:
                self.tabela.column(col, width=0, stretch=False, minwidth=0)
        
        # Salvar configuração
        self.salvar_configuracao_colunas(colunas_selecionadas)
        
        self.status_bar.config(text=f"Colunas visíveis: {len(colunas_selecionadas)} de {len(self.colunas)}")
    
    def _get_minwidth_coluna(self, col):
        """Retorna o minwidth padrão para cada coluna"""
        if 'DESCRI' in col:
            return 300
        elif col == 'CLIENTE':
            return 150
        elif col in ('ARQUIVO_ORIGEM',):
            return 150
        elif col in ('CODE', 'ITEM'):
            return 100
        elif col == 'IMAGEM':
            return 60
        else:
            return 120
    
    def _get_largura_coluna(self, col):
        """Retorna a largura padrão para cada coluna"""
        if 'DESCRI' in col:
            return 800
        elif col == 'CLIENTE':
            return 200
        elif col in ('ARQUIVO_ORIGEM',):
            return 250
        elif col in ('CODE', 'ITEM'):
            return 150
        elif col == 'IMAGEM':
            return 80
        else:
            return 200
    
    def salvar_configuracao_colunas(self, colunas_selecionadas):
        """Salva as colunas visíveis em arquivo JSON"""
        arquivo_config = os.path.join(DATA_DIR, "config_colunas.json")
        try:
            with open(arquivo_config, 'w', encoding='utf-8') as f:
                json.dump({'colunas_visiveis': colunas_selecionadas}, f, indent=4)
        except Exception as e:
            print(f"DEBUG: Erro ao salvar configuração de colunas: {e}")
    
    def carregar_configuracao_colunas(self):
        """Carrega as colunas visíveis do arquivo JSON"""
        arquivo_config = os.path.join(DATA_DIR, "config_colunas.json")
        if os.path.exists(arquivo_config):
            try:
                with open(arquivo_config, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    return config.get('colunas_visiveis', None)
            except Exception as e:
                print(f"DEBUG: Erro ao carregar configuração de colunas: {e}")
        return None
    
    def atualizar_lista_clientes(self):
        """Atualiza a lista de clientes no combobox"""
        clientes = ["Todos"] + listar_clientes()
        self.combo_clientes['values'] = clientes
    
    def executar_busca(self):
        """Executa a busca e atualiza a tabela"""
        termo = self.termo_busca.get().strip()
        cliente = self.cliente_selecionado.get()
        
        # Limpar tabela atual
        for item in self.tabela.get_children():
            self.tabela.delete(item)
        
        # Se não tiver termo mas tiver cliente selecionado, busca por cliente
        if not termo and cliente != "Todos":
            termo = "*"  # Busca tudo para este cliente
        
        # Se não tiver termo E cliente for "Todos", mostra todos os produtos
        if not termo and cliente == "Todos":
            termo = "*"  # Busca todos os produtos
        
        try:
            resultados = buscar_produtos(termo, cliente)
            
            # Calcular total da coluna TOTAL CTNS
            total_ctns = 0
            idx_total_ctns = None
            colunas_list = list(self.colunas)
            if 'TOTAL CTNS' in colunas_list:
                idx_total_ctns = colunas_list.index('TOTAL CTNS')
            
            # Calcular total da coluna TOTAL GROSS WEIGHT (KG)
            total_gross = 0.0
            idx_gross = None
            if 'TOTAL GROSS WEIGHT( kg )' in colunas_list:
                idx_gross = colunas_list.index('TOTAL GROSS WEIGHT( kg )')
            
            for resultado in resultados:
                self.tabela.insert('', 'end', values=resultado)
                # Somar CTNS se a coluna existir
                if idx_total_ctns is not None and idx_total_ctns < len(resultado):
                    try:
                        val = str(resultado[idx_total_ctns]).replace('.', '').replace(',', '.')
                        if val:
                            total_ctns += float(val)
                    except:
                        pass
                # Somar GROSS WEIGHT se a coluna existir
                if idx_gross is not None and idx_gross < len(resultado):
                    try:
                        val = str(resultado[idx_gross]).replace('.', '').replace(',', '.')
                        if val:
                            total_gross += float(val)
                    except:
                        pass
            
            # Mostrar resultados + totais
            texto_resultado = f"{len(resultados)} resultados"
            if total_ctns > 0:
                texto_resultado += f" | Total CTNS: {int(total_ctns)}"
            if total_gross > 0:
                texto_resultado += f" | Total Gross: {total_gross:.2f}"
            self.label_resultados.config(text=texto_resultado)
            self.status_bar.config(text=f"Busca concluída: {len(resultados)} resultados")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro na busca: {str(e)}")
    
    def limpar_busca(self):
        """Limpa os campos de busca"""
        self.termo_busca.set("")
        self.cliente_selecionado.set("Todos")
        
        for item in self.tabela.get_children():
            self.tabela.delete(item)
        
        self.label_resultados.config(text="0 resultados encontrados")
        self.status_bar.config(text="Busca limpa")
    
    def importar_pasta(self):
        """Importa planilhas da pasta planilhas/ em thread separada"""
        def importar_thread():
            try:
                self.status_bar.config(text="Importando planilhas da pasta 'planilhas/'...")
                self.janela.update()
                
                total = importar_todas_planilhas(self.atualizar_progresso)
                
                self.atualizar_estatisticas()
                self.atualizar_lista_clientes()
                
                messagebox.showinfo("Importação Concluída", f"Foram importados {total} produtos com sucesso!")
                self.status_bar.config(text=f"Importação concluída: {total} produtos")
                
            except Exception as e:
                messagebox.showerror("Erro na Importação", f"Ocorreu um erro: {str(e)}")
                self.status_bar.config(text="Erro na importação")
        
        threading.Thread(target=importar_thread, daemon=True).start()
    
    def selecionar_arquivos(self):
        """Permite selecionar múltiplos arquivos para importar"""
        arquivos = filedialog.askopenfilenames(
            title="Selecione as planilhas para importar",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls"), ("Todos os arquivos", "*.*")]
        )
        
        if arquivos:
            def importar_thread():
                try:
                    self.status_bar.config(text=f"Importando {len(arquivos)} arquivos...")
                    self.janela.update()
                    
                    total = importar_arquivos_selecionados(arquivos, self.atualizar_progresso)
                    
                    self.atualizar_estatisticas()
                    self.atualizar_lista_clientes()
                    
                    messagebox.showinfo("Importação Concluída", f"Foram importados {total} produtos com sucesso!")
                    self.status_bar.config(text=f"Importação concluída: {total} produtos")
                    
                except Exception as e:
                    messagebox.showerror("Erro na Importação", f"Ocorreu um erro: {str(e)}")
                    self.status_bar.config(text="Erro na importação")
            
            threading.Thread(target=importar_thread, daemon=True).start()
    
    def limpar_banco(self):
        """Limpa todos os dados do banco"""
        if messagebox.askyesno("Confirmar Limpeza", "Tem certeza que deseja limpar todos os dados? Esta ação não pode ser desfeita!"):
            try:
                cursor = get_cursor()
                cursor.execute("DELETE FROM produtos")
                cursor.execute("DELETE FROM importacoes")
                get_connection().commit()
                
                self.atualizar_estatisticas()
                self.atualizar_lista_clientes()
                self.limpar_busca()
                
                messagebox.showinfo("Banco Limpo", "Todos os dados foram removidos com sucesso!")
                self.status_bar.config(text="Banco de dados limpo")
                
            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao limpar banco: {str(e)}")
    
    def exportar_excel(self):
        """Exporta resultados para Excel - apenas planilha selecionada no filtro"""
        resultados = []
        for item in self.tabela.get_children():
            resultados.append(self.tabela.item(item)['values'])
        
        if not resultados:
            messagebox.showwarning("Aviso", "Não há resultados para exportar!")
            return
        
        print(f"DEBUG EXPORT PY UI: inicio exportar_excel | linhas={len(resultados)}")
        print(f"DEBUG EXPORT PY UI: pasta_configurada={self.pasta_exportacoes}")
        arquivo = exportar_resultados(resultados, 'excel', self.pasta_exportacoes, colunas=list(self.colunas))
        
        if arquivo:
            if isinstance(arquivo, list):
                messagebox.showinfo("Exportação Concluída", f"Exportados {len(arquivo)} arquivos em:\n{self.pasta_exportacoes}")
                self.status_bar.config(text=f"Exportados {len(arquivo)} arquivos")
            else:
                messagebox.showinfo("Exportação Concluída", f"Resultados exportados para:\n{arquivo}")
                self.status_bar.config(text=f"Exportado para {arquivo}")
        else:
            messagebox.showerror("Erro de Exportação", "A exportação não gerou arquivo. Veja os logs DEBUG no terminal.")
    
        
        
    def ver_foto(self):
        """Abre janela para visualizar foto do produto selecionado"""
        selecionado = self.tabela.selection()
        if not selecionado:
            messagebox.showwarning("Aviso", "Selecione um produto na tabela para ver a foto.")
            return
        
        item = self.tabela.item(selecionado[0])
        valores = item['values']
        
        if len(valores) < 8:
            messagebox.showwarning("Aviso", "Produto não possui informações completas.")
            return
        
        cliente = valores[0]
        
        # Encontrar índice da coluna PICTURE
        colunas_banco = list(get_colunas_banco().keys())
        picture_idx = None
        for i, col in enumerate(colunas_banco):
            if col == 'PICTURE':
                picture_idx = i
                break
        
        if picture_idx is None or picture_idx >= len(valores):
            messagebox.showwarning("Aviso", "Coluna PICTURE não encontrada.")
            return
            
        picture_valor = valores[picture_idx]
        
        # Debug para mostrar PICTURE e caminho
        print(f"DEBUG: Valor da coluna PICTURE: '{picture_valor}'")
        print(f"DEBUG: Cliente: '{cliente}'")
        print(f"DEBUG: Índice PICTURE: {picture_idx}")
        print(f"DEBUG: Colunas banco: {colunas_banco[:10]}...")
        
        # Se PICTURE tiver valor, usar como nome da imagem
        caminho_imagem_custom = None
        if picture_valor is not None and picture_valor != '':
            # Converter para string se for número
            picture_str = str(picture_valor).strip()
            if picture_str:
                # Se for URL, abre no navegador e encerra fluxo local
                if picture_str.lower().startswith(('http://', 'https://')):
                    try:
                        import webbrowser
                        if messagebox.askyesno(
                            "Imagem por URL",
                            f"Campo PICTURE contém URL:\n{picture_str}\n\nDeseja abrir no navegador?"
                        ):
                            webbrowser.open(picture_str)
                    except Exception as e:
                        messagebox.showerror("Erro", f"Não foi possível abrir URL da imagem:\n{str(e)}")
                    return
                
                # Se já vier com extensão/caminho, respeita o valor salvo
                if picture_str.lower().endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp', '.webp')):
                    nome_imagem = os.path.basename(picture_str)
                    if os.path.isabs(picture_str) or ('/' in picture_str) or ('\\' in picture_str):
                        caminho_imagem_custom = picture_str
                else:
                    nome_imagem = f"{picture_str}.jpg"
            else:
                nome_imagem = "default.jpg"
        else:
            # Fallback: usar CODE
            code_idx = None
            for i, col in enumerate(colunas_banco):
                if col == 'CODE':
                    code_idx = i
                    break
            if code_idx is not None and code_idx < len(valores):
                codigo = valores[code_idx]
                nome_imagem = f"{codigo}.jpg" if codigo else "default.jpg"
            else:
                nome_imagem = "default.jpg"
        
        caminho_imagem = caminho_imagem_custom or os.path.join(IMAGENS_DIR, cliente, nome_imagem)
        
        print(f"DEBUG: Caminho da imagem: {caminho_imagem}")
        print(f"DEBUG: Imagem existe? {os.path.exists(caminho_imagem)}")
        
        # Criar pasta se não existir
        pasta_imagens = os.path.join(IMAGENS_DIR, cliente)
        if not os.path.exists(pasta_imagens):
            os.makedirs(pasta_imagens)
        
        # Janela para visualizar imagem
        janela_foto = tk.Toplevel(self.janela)
        janela_foto.title(f"🖼️ Foto do Produto - {nome_imagem.replace('.jpg', '')}")
        janela_foto.geometry("600x500")
        janela_foto.configure(bg='#2c3e50')
        
        # Frame principal
        frame_principal = tk.Frame(janela_foto, bg='#2c3e50')
        frame_principal.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Informações do produto
        info_text = f"Cliente: {cliente}\nImagem: {nome_imagem}\nDescrição: {valores[3] if len(valores) > 3 else 'N/A'}"
        tk.Label(frame_principal, text=info_text, font=('Arial', 12, 'bold'), 
                bg='#2c3e50', fg='white', justify='left').pack(pady=(0, 10))
        
        try:
            if os.path.exists(caminho_imagem):
                # Carregar e exibir imagem
                img = Image.open(caminho_imagem)
                img.thumbnail((500, 300), Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                
                label_img = tk.Label(frame_principal, image=photo, bg='#2c3e50')
                label_img.image = photo  # Manter referência
                label_img.pack(pady=10)
                
                status = "✅ Foto encontrada"
                cor = '#2ecc71'
            else:
                # Imagem não encontrada
                label_img = tk.Label(frame_principal, text="📷\n\nFOTO NÃO ENCONTRADA\n\nAdicione a imagem em:\n" + caminho_imagem,
                                 font=('Arial', 14), bg='#34495e', fg='#e74c3c', 
                                 width=30, height=15, relief='raised', bd=2)
                label_img.pack(pady=10)
                
                status = "❌ Foto não encontrada"
                cor = '#e74c3c'
            
            tk.Label(frame_principal, text=status, font=('Arial', 12, 'bold'), 
                    bg='#2c3e50', fg=cor).pack(pady=10)
            
        except Exception as e:
            tk.Label(frame_principal, text=f"Erro ao carregar imagem: {str(e)}", 
                    font=('Arial', 11), bg='#2c3e50', fg='#e74c3c').pack(pady=10)
        
        # Botões
        frame_botoes = tk.Frame(janela_foto, bg='#2c3e50')
        frame_botoes.pack(pady=10)
        
        tk.Button(frame_botoes, text="📁 Abrir Pasta", 
                 command=lambda: os.startfile(os.path.join(IMAGENS_DIR, cliente)),
                 bg='#3498db', fg='white', font=('Arial', 10, 'bold'), 
                 width=15, height=1).pack(side='left', padx=5)
        
        tk.Button(frame_botoes, text="✅ Fechar", command=janela_foto.destroy,
                 bg='#27ae60', fg='white', font=('Arial', 10, 'bold'), 
                 width=15, height=1).pack(side='left', padx=5)
        
        # Centralizar janela
        janela_foto.transient(self.janela)
        janela_foto.grab_set()
    
    def exportar_csv(self):
        """Exporta resultados para CSV"""
        resultados = []
        for item in self.tabela.get_children():
            resultados.append(self.tabela.item(item)['values'])
        
        if not resultados:
            messagebox.showwarning("Aviso", "Não há resultados para exportar!")
            return
        
        print(f"DEBUG EXPORT PY UI: inicio exportar_csv | linhas={len(resultados)}")
        print(f"DEBUG EXPORT PY UI: pasta_configurada={self.pasta_exportacoes}")
        arquivo = exportar_resultados(resultados, 'csv', self.pasta_exportacoes, colunas=list(self.colunas))
        if arquivo:
            if isinstance(arquivo, list):
                messagebox.showinfo("Exportação Concluída", f"Exportados {len(arquivo)} arquivos em:\n{self.pasta_exportacoes}")
                self.status_bar.config(text=f"Exportados {len(arquivo)} arquivos")
            else:
                messagebox.showinfo("Exportação Concluída", f"Resultados exportados para:\n{arquivo}")
                self.status_bar.config(text=f"Exportado para {arquivo}")
        else:
            messagebox.showerror("Erro de Exportação", "A exportação não gerou arquivo. Veja os logs DEBUG no terminal.")
    
    def atualizar_progresso(self, mensagem):
        """Atualiza a barra de status com mensagens de progresso"""
        self.status_bar.config(text=mensagem)
        self.janela.update()
    
    def mostrar_capacidade(self):
        """Mostra janela bonita com capacidade do sistema"""
        janela_capacidade = tk.Toplevel(self.janela)
        janela_capacidade.title("📊 Capacidade do Sistema")
        janela_capacidade.geometry("750x650")
        janela_capacidade.configure(bg='#2c3e50')
        janela_capacidade.resizable(False, False)
        
        # Título principal
        titulo = tk.Label(janela_capacidade, text="🚀 CAPACIDADE MÁXIMA DO SISTEMA", 
                        font=('Arial', 18, 'bold'), bg='#2c3e50', fg='#3498db')
        titulo.pack(pady=15)
        
        # Frame principal com scroll
        frame_principal_container = tk.Frame(janela_capacidade, bg='#2c3e50')
        frame_principal_container.pack(fill='both', expand=True, padx=20, pady=(0, 10))
        
        # Canvas para scroll
        canvas = tk.Canvas(frame_principal_container, bg='#2c3e50', highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame_principal_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#34495e', relief='raised', bd=2)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Estatísticas atuais
        frame_stats = tk.LabelFrame(scrollable_frame, text="📊 ESTATÍSTICAS ATUAIS", 
                                  font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_stats.pack(fill='x', padx=15, pady=15)
        
        total_produtos = contar_produtos()
        total_importacoes = contar_importacoes()
        
        stats_text = f"""
📦 Produtos cadastrados: {total_produtos:,}
📁 Importações realizadas: {total_importacoes:,}
🗄️ Banco de dados: SQLite otimizado
⚡ Índices ativos: 5 índices de busca
        """
        
        label_stats = tk.Label(frame_stats, text=stats_text, font=('Arial', 12), 
                              bg='#34495e', fg='#ecf0f1', justify='left')
        label_stats.pack(padx=20, pady=15)
        
        # Capacidade máxima
        frame_capacidade = tk.LabelFrame(scrollable_frame, text="⚡ CAPACIDADE MÁXIMA", 
                                       font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_capacidade.pack(fill='x', padx=15, pady=15)
        
        capacidade_text = """
📁 PLANILHAS MÁXIMAS:
   • Simultâneas: 500+ arquivos
   • Por planilha: Até 100.000 linhas
   • Total combinado: 1.000.000+ produtos

🗄️ BANCO DE DADOS:
   • Capacidade teórica: 140TB
   • Recomendado: Até 10GB
   • Seu sistema: Fácilmente 1-2GB

⚡ PERFORMANCE:
   • Busca: <1 segundo (mesmo com 1M produtos)
   • Importação: 30-120 segundos
   • Memória RAM: 2GB mínimo, 4GB+ ideal

💾 RECURSOS AVANÇADOS:
   • Índices otimizados para busca ultra-rápida
   • Cache inteligente para performance máxima
   • Thread-safe para importação sem travar
   • Arquitetura escalável para grandes volumes

🚀 VERSÃO PLUS DISPONÍVEL:
   • CAPACIDADE EXPANDIDA: 5 MILHÕES de produtos
   • Planilhas simultâneas: 2.000+ arquivos
   • Interface dark profissional
   • Otimizações de última geração
        """
        
        label_capacidade = tk.Label(frame_capacidade, text=capacidade_text, 
                                   font=('Arial', 11), bg='#34495e', fg='#ecf0f1', justify='left')
        label_capacidade.pack(padx=20, pady=15)
        
        # Frame de comparação
        frame_comparacao = tk.LabelFrame(scrollable_frame, text="🏆 COMPARAÇÃO DE PERFORMANCE", 
                                        font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_comparacao.pack(fill='x', padx=15, pady=15)
        
        comparacao_text = """
📊 PERFORMANCE COMPARATIVA:
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   │ Sistema   │ Capacidade    │ Planilhas │ Performance │
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   │ Normal    │ 1 Milhão     │ 500+     │ Ultra-rápida │
   │ PLUS      │ 5 Milhões    │ 2.000+   │ Extrema      │
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

⏱️ TEMPOS DE IMPORTAÇÃO:
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   │ Volume    │ Tempo Normal │ Tempo PLUS │ Memória   │
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
   │ 50 plan   │ ~10s        │ ~5s       │ ~50MB     │
   │ 200 plan  │ ~45s        │ ~20s      │ ~150MB    │
   │ 500 plan  │ ~2min       │ ~1min     │ ~300MB    │
   │ 1000 plan │ N/A         │ ~2min     │ ~500MB    │
   ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
        """
        
        label_comparacao = tk.Label(frame_comparacao, text=comparacao_text, 
                                   font=('Courier', 10), bg='#34495e', fg='#2ecc71', justify='left')
        label_comparacao.pack(padx=20, pady=15)
        
        # Frame de recomendações
        frame_recomendacoes = tk.LabelFrame(scrollable_frame, text="💡 RECOMENDAÇÕES", 
                                          font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_recomendacoes.pack(fill='x', padx=15, pady=15)
        
        recomendacoes_text = """
✅ DICAS PARA USAR MELHOR:
   • Importe suas planilhas em pequenos grupos (até 100 por vez)
   • Faça cópia do arquivo 'banco.db' regularmente
   • Use nomes claros para seus arquivos Excel
   • Mantenha suas planilhas organizadas

🚀 PRECISA DE MAIS CAPACIDADE?
   • Sistema PLUS suporta 5 MILHÕES de produtos
   • Trabalha com 2.000+ planilhas ao mesmo tempo
   • Interface super profissional e moderna
   • Performance extremamente rápida
   
💎 Para usar a versão PLUS: Execute o sistema avançado 

� DICAS IMPORTANTES:
   • Salve seu trabalho antes de fechar o sistema
   • Mantenha backup das planilhas originais
   • Use buscas específicas para encontrar mais rápido
   • Exporte resultados quando precisar compartilhar

🎯 USO IDEAL DO SISTEMA:
   • Empresas pequenas/médias: Sistema normal
   • Grandes corporações: Sistema PLUS
   • Economia de tempo: 90% mais rápido que busca manual
   • Crescimento: Sistema cresce com seu negócio
        """
        
        label_recomendacoes = tk.Label(frame_recomendacoes, text=recomendacoes_text, 
                                       font=('Arial', 11), bg='#34495e', fg='#f39c12', justify='left')
        label_recomendacoes.pack(padx=20, pady=15)
        
        # Pack do canvas e scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Botão fechar
        tk.Button(janela_capacidade, text="✅ Entendido", command=janela_capacidade.destroy,
                 bg='#27ae60', fg='white', font=('Arial', 12, 'bold'), width=15, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(pady=15)
        
        # Centralizar janela
        janela_capacidade.transient(self.janela)
        janela_capacidade.grab_set()
        
        # Habilitar scroll com mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind("<MouseWheel>", _on_mousewheel)
    
    def abrir_manual(self):
        """Abre o manual do usuário em container igual ao de capacidade"""
        # Sempre abre no container do sistema (ignora arquivo externo)
        self.mostrar_manual_em_janela()
    
    def mostrar_manual_em_janela(self):
        """Mostra o manual do usuário em uma janela do sistema"""
        janela_manual = tk.Toplevel(self.janela)
        janela_manual.title("📖 Manual do Usuário")
        janela_manual.geometry("750x650")
        janela_manual.configure(bg='#2c3e50')
        janela_manual.resizable(False, False)
        
        # Título principal
        titulo = tk.Label(janela_manual, text="📖 MANUAL DO USUÁRIO", 
                        font=('Arial', 18, 'bold'), bg='#2c3e50', fg='#3498db')
        titulo.pack(pady=15)
        
        # Frame principal com scroll (igual ao de capacidade)
        frame_principal_container = tk.Frame(janela_manual, bg='#2c3e50')
        frame_principal_container.pack(fill='both', expand=True, padx=20, pady=(0, 10))
        
        # Canvas para scroll
        canvas = tk.Canvas(frame_principal_container, bg='#2c3e50', highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame_principal_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#34495e', relief='raised', bd=2)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Seções do manual
        frame_inicio = tk.LabelFrame(scrollable_frame, text="🚀 COMEÇANDO (3 PASSOS)", 
                                    font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_inicio.pack(fill='x', padx=15, pady=15)
        
        texto_inicio = """
1️⃣ Coloque suas planilhas .xlsx/.xls na pasta "planilhas"
2️⃣ Clique em "📁 Importar Pasta 'planilhas/'"
3️⃣ Busque qualquer produto no campo "Buscar"
        """
        
        label_inicio = tk.Label(frame_inicio, text=texto_inicio, font=('Arial', 12), 
                              bg='#34495e', fg='#ecf0f1', justify='left')
        label_inicio.pack(padx=20, pady=15)
        
        # Botões de busca
        frame_busca = tk.LabelFrame(scrollable_frame, text="🔍 BOTÕES DE BUSCA", 
                                  font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_busca.pack(fill='x', padx=15, pady=15)
        
        texto_busca = """
🔎 BUSCAR: Procura em todos os produtos
🗑️ LIMPAR: Limpa busca e resultados
📄 CLIENTE: Filtra por cliente específico
        """
        
        label_busca = tk.Label(frame_busca, text=texto_busca, font=('Arial', 12), 
                              bg='#34495e', fg='#ecf0f1', justify='left')
        label_busca.pack(padx=20, pady=15)
        
        # Botões de importação
        frame_import = tk.LabelFrame(scrollable_frame, text="📂 BOTÕES DE IMPORTAÇÃO", 
                                   font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_import.pack(fill='x', padx=15, pady=15)
        
        texto_import = """
📁 IMPORTAR PASTA: Importa todas as planilhas da pasta
📄 SELECIONAR ARQUIVOS: Escolhe planilhas específicas
🔄 LIMPAR BANCO: Apaga TODOS os dados (cuidado!)
📊 CAPACIDADE: Mostra estatísticas e limites
        """
        
        label_import = tk.Label(frame_import, text=texto_import, font=('Arial', 12), 
                               bg='#34495e', fg='#ecf0f1', justify='left')
        label_import.pack(padx=20, pady=15)
        
        # Como funciona
        frame_funciona = tk.LabelFrame(scrollable_frame, text="⚡ COMO FUNCIONA", 
                                      font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_funciona.pack(fill='x', padx=15, pady=15)
        
        texto_funciona = """
• Detecta colunas automaticamente (código, descrição, peso, valor, NCM)
• Salva em banco de dados ultra-rápido
• Busca instantânea (< 1 segundo)
• Funciona 100% offline
        """
        
        label_funciona = tk.Label(frame_funciona, text=texto_funciona, font=('Arial', 12), 
                                 bg='#34495e', fg='#ecf0f1', justify='left')
        label_funciona.pack(padx=20, pady=15)
        
        # Dicas
        frame_dicas = tk.LabelFrame(scrollable_frame, text="🎯 DICAS IMPORTANTES", 
                                   font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_dicas.pack(fill='x', padx=15, pady=15)
        
        texto_dicas = """
• Use termos curtos: "parafuso" (melhor que frases longas)
• Busque por código: "85044030" (para NCM)
• Filtre por cliente para resultados mais precisos
• Exporte resultados para compartilhar
        """
        
        label_dicas = tk.Label(frame_dicas, text=texto_dicas, font=('Arial', 12), 
                              bg='#34495e', fg='#f39c12', justify='left')
        label_dicas.pack(padx=20, pady=15)
        
        # Suporte
        frame_suporte = tk.LabelFrame(scrollable_frame, text="🆘 SUPORTE", 
                                     font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_suporte.pack(fill='x', padx=15, pady=15)
        
        texto_suporte = """
• Erro "nenhuma planilha": verifique pasta "planilhas"
• Busca vazia: importe as planilhas primeiro
• Erro na importação: feche o arquivo Excel
• Sistema PLUS: Para mais capacidade, use a versão avançada
        """
        
        label_suporte = tk.Label(frame_suporte, text=texto_suporte, font=('Arial', 12), 
                                 bg='#34495e', fg='#2ecc71', justify='left')
        label_suporte.pack(padx=20, pady=15)
        
        # Pack do canvas e scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Botão fechar
        tk.Button(janela_manual, text="✅ Entendido", command=janela_manual.destroy,
                 bg='#27ae60', fg='white', font=('Arial', 12, 'bold'), width=15, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(pady=15)
        
        # Centralizar janela
        janela_manual.transient(self.janela)
        janela_manual.grab_set()
        
        # Habilitar scroll com mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind("<MouseWheel>", _on_mousewheel)
    
    def abrir_sistema_plus(self):
        """Abre o sistema PLUS com informações"""
        janela_plus = tk.Toplevel(self.janela)
        janela_plus.title("🚀 Sistema PLUS")
        janela_plus.geometry("700x500")
        janela_plus.configure(bg='#2c3e50')
        janela_plus.resizable(False, False)
        
        # Título
        titulo = tk.Label(janela_plus, text="🚀 SISTEMA PLUS - VERSÃO AVANÇADA", 
                        font=('Arial', 18, 'bold'), bg='#2c3e50', fg='#e74c3c')
        titulo.pack(pady=20)
        
        # Frame principal com scroll
        frame_principal_container = tk.Frame(janela_plus, bg='#2c3e50')
        frame_principal_container.pack(fill='both', expand=True, padx=20, pady=(0, 10))
        
        # Canvas para scroll
        canvas = tk.Canvas(frame_principal_container, bg='#2c3e50', highlightthickness=0)
        scrollbar = ttk.Scrollbar(frame_principal_container, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#34495e', relief='raised', bd=2)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Benefícios do PLUS
        frame_beneficios = tk.LabelFrame(scrollable_frame, text="💎 BENEFÍCIOS EXCLUSIVOS", 
                                        font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_beneficios.pack(fill='x', padx=15, pady=15)
        
        beneficios_text = """
📦 CAPACIDADE EXPANDIDA:
   • 5 MILHÕES de produtos (vs 1 milhão)
   • 2.000+ planilhas simultâneas (vs 500)
   • Importação 10x mais rápida
   
⚡ PERFORMANCE EXTREMA:
   • Busca em tempo real (< 0.1 segundo)
   • Interface dark profissional
   • Otimizações avançadas
   
🎯 PARA GRANDES EMPRESAS:
   • Múltiplos departamentos
   • Altíssimo volume de dados
   • Nível corporativo
        """
        
        label_beneficios = tk.Label(frame_beneficios, text=beneficios_text, 
                                   font=('Arial', 12), bg='#34495e', fg='#ecf0f1', justify='left')
        label_beneficios.pack(padx=20, pady=15)
        
        # Como acessar
        frame_acesso = tk.LabelFrame(scrollable_frame, text="🔑 COMO ACESSAR", 
                                    font=('Arial', 14, 'bold'), bg='#34495e', fg='white')
        frame_acesso.pack(fill='x', padx=15, pady=15)
        
        acesso_text = """
📋 REQUISITOS:
   • Computador com 4GB+ de RAM
   • Espaço em disco para dados
   • Windows 10 ou superior
   
🚀 PARA USAR O SISTEMA PLUS:
   1️⃣ Clique no botão "🚀 Acessar Sistema PLUS" abaixo
   2️⃣ O sistema avançado será aberto automaticamente
   3️⃣ Aproveite a capacidade máxima!
   
💡 DICAS:
   • Importe dados em lotes menores
   • Faça backup regularmente
   • Use buscas específicas
        """
        
        label_acesso = tk.Label(frame_acesso, text=acesso_text, 
                              font=('Arial', 12), bg='#34495e', fg='#f39c12', justify='left')
        label_acesso.pack(padx=20, pady=15)
        
        # Pack do canvas e scrollbar
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Botões
        frame_botoes = tk.Frame(janela_plus, bg='#2c3e50')
        frame_botoes.pack(pady=15)
        
        tk.Button(frame_botoes, text="� Acessar Sistema PLUS", command=self.executar_sistema_plus,
                 bg='#e74c3c', fg='white', font=('Arial', 12, 'bold'), width=20, height=2,
                 relief='raised', bd=3, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_botoes, text="� Ver Manual PLUS", command=self.abrir_manual,
                 bg='#9b59b6', fg='white', font=('Arial', 10, 'bold'), width=20, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        tk.Button(frame_botoes, text="✅ Entendido", command=janela_plus.destroy,
                 bg='#27ae60', fg='white', font=('Arial', 12, 'bold'), width=15, height=2,
                 relief='raised', bd=2, cursor='hand2').pack(side='left', padx=5)
        
        # Centralizar janela
        janela_plus.transient(self.janela)
        janela_plus.grab_set()
        
        # Habilitar scroll com mouse
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind("<MouseWheel>", _on_mousewheel)
    
    def executar_sistema_plus(self):
        """Executa o sistema PLUS automaticamente"""
        try:
            import subprocess
            import sys
            
            # Fecha o sistema atual
            self.janela.destroy()
            
            # Executa o sistema PLUS
            subprocess.Popen([sys.executable, "sistema_plus.py"])
            
        except Exception as e:
            messagebox.showerror("Erro", 
                "Não foi possível abrir o Sistema PLUS. Verifique se o arquivo 'sistema_plus.py' existe na pasta.")
    
    def ir_para_sistema_plus(self):
        """Navega para o Sistema Plus"""
        if messagebox.askyesno("Navegar para Sistema PLUS", 
                              "Deseja fechar o Sistema Original e abrir o Sistema PLUS?"):
            self.executar_sistema_plus()
    
    def executar(self):
        """Inicia a interface"""
        self.janela.mainloop()

# ==============================
# INICIALIZAÇÃO
# ==============================

if __name__ == "__main__":
    # NÃO limpar banco - acumular planilhas
    # limpar_banco_dados()
    
    app = SistemaPlanilhas()
    app.executar()
