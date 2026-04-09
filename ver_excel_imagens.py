import os
import zipfile
from openpyxl import load_workbook

def analisar_excel_para_imagens(caminho_arquivo):
    """Analisa um arquivo Excel para encontrar imagens"""
    print(f"=== ANALISANDO: {caminho_arquivo} ===")
    
    # Método 1: Verificar estrutura ZIP
    print("\n1. Verificando estrutura ZIP do Excel:")
    try:
        with zipfile.ZipFile(caminho_arquivo, 'r') as zip_ref:
            todos_arquivos = zip_ref.namelist()
            print(f"Total de arquivos no ZIP: {len(todos_arquivos)}")
            
            # Procurar por imagens
            arquivos_imagem = [f for f in todos_arquivos if f.lower().endswith(('.png', '.jpg', '.jpeg', '.gif', '.bmp'))]
            print(f"Arquivos de imagem: {len(arquivos_imagem)}")
            
            if arquivos_imagem:
                print("Imagens encontradas:")
                for img in arquivos_imagem:
                    print(f"  - {img}")
            else:
                print("Nenhuma imagem encontrada no ZIP")
                
            # Mostrar estrutura de pastas
            pastas = set()
            for f in todos_arquivos:
                if '/' in f:
                    pastas.add(f.split('/')[0])
            print(f"Pastas encontradas: {sorted(pastas)}")
            
    except Exception as e:
        print(f"Erro ao ler ZIP: {e}")
    
    # Método 2: Verificar com openpyxl
    print("\n2. Verificando com openpyxl:")
    try:
        wb = load_workbook(caminho_arquivo)
        ws = wb.active
        
        # Verificar se há imagens na worksheet
        if hasattr(ws, '_images'):
            print(f"Imagens na worksheet: {len(ws._images)}")
            for i, img in enumerate(ws._images):
                print(f"  Imagem {i+1}: {img}")
        else:
            print("Nenhuma imagem encontrada na worksheet")
            
        # Verificar valores da coluna PICTURE
        print("\n3. Verificando coluna PICTURE:")
        for row_num in range(1, min(10, ws.max_row + 1)):
            row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
            if len(row) > 0:
                picture_val = row[0]  # Primeira coluna
                print(f"  Linha {row_num}: PICTURE = '{picture_val}'")
                
    except Exception as e:
        print(f"Erro ao ler com openpyxl: {e}")

if __name__ == "__main__":
    # Testar com o arquivo
    arquivo_excel = "E:/planilhas-teste/05-03-2026 - 25-608_ESTUDO_250503 - COM LI ANP.xlsx"
    
    if os.path.exists(arquivo_excel):
        analisar_excel_para_imagens(arquivo_excel)
    else:
        print(f"Arquivo não encontrado: {arquivo_excel}")
        
        # Listar arquivos disponíveis
        pasta = "E:/planilhas-teste"
        if os.path.exists(pasta):
            print(f"\nArquivos disponíveis em {pasta}:")
            for f in os.listdir(pasta):
                if f.endswith(('.xlsx', '.xls')):
                    print(f"  - {f}")
