#!/usr/bin/env python3
from openpyxl import load_workbook
import os

print("=== ANALISANDO EXCEL ===")
wb = load_workbook('E:/planilhas-teste/celio.planilha.xlsx', read_only=True)
ws = wb.active

print(f"Total de linhas: {ws.max_row}")
print(f"Total de colunas: {ws.max_column}")

print("\n=== PRIMEIRAS 5 LINHAS ===")
for row_num in range(1, 6):
    row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    cells = [c for c in row if c]
    print(f"\nLinha {row_num}: {len(cells)} células preenchidas")
    for i, cell in enumerate(row[:20]):
        if cell:
            print(f"  [{i}] = {repr(str(cell)[:50])}")

# Encontrar linha com mais cabeçalhos
print("\n=== PROCURANDO CABEÇALHOS ===")
max_celulas = 0
linha_cabecalho = 1
cabecalhos = []

for row_num in range(1, min(21, ws.max_row + 1)):
    row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    celulas = [c for c in row if c and str(c).strip()]
    if len(celulas) > max_celulas:
        max_celulas = len(celulas)
        cabecalhos = [str(c).strip() for c in row if c and str(c).strip()]
        linha_cabecalho = row_num

print(f"Cabeçalhos encontrados na linha {linha_cabecalho}")
print(f"Total de colunas: {len(cabecalhos)}")
print(f"Primeiros 10: {cabecalhos[:10]}")

# Contar linhas de dados
print(f"\n=== CONTANDO LINHAS DE DADOS ===")
count = 0
for row in ws.iter_rows(min_row=linha_cabecalho+1, values_only=True):
    if any(c for c in row if c):
        count += 1

print(f"Linhas de dados: {count}")

wb.close()
print("\n=== FIM ===")
os.system("pause")
