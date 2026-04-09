import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image
import shutil
from openpyxl import load_workbook

class UploadImagens:
    def __init__(self):
        self.janela = tk.Tk()
        self.janela.title("📸 Upload Automático de Imagens")
        self.janela.geometry("800x600")
        self.janela.configure(bg='#2c3e50')
        
        self.cliente_selecionado = None
        self.arquivo_excel = None
        self.imagens_selecionadas = []
        
        self.criar_interface()
        
    def criar_interface(self):
        # Frame principal
        frame = tk.Frame(self.janela, bg='#2c3e50')
        frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        # Título
        tk.Label(frame, text="📸 UPLOAD AUTOMÁTICO DE IMAGENS", 
                font=('Arial', 16, 'bold'), bg='#2c3e50', fg='white').pack(pady=(0, 20))
        
        # Selecionar arquivo Excel
        frame_excel = tk.Frame(frame, bg='#34495e')
        frame_excel.pack(fill='x', pady=10)
        
        tk.Label(frame_excel, text="📄 Arquivo Excel:", font=('Arial', 12), 
                bg='#34495e', fg='white').pack(side='left', padx=10, pady=10)
        
        self.btn_excel = tk.Button(frame_excel, text="Selecionar Excel", 
                                command=self.selecionar_excel,
                                bg='#3498db', fg='white', font=('Arial', 10, 'bold'),
                                padx=20, pady=10)
        self.btn_excel.pack(side='left', padx=10, pady=10)
        
        self.lbl_excel = tk.Label(frame_excel, text="Nenhum arquivo selecionado", 
                                font=('Arial', 10), bg='#34495e', fg='#ecf0f1')
        self.lbl_excel.pack(side='left', padx=10, pady=10)
        
        # Selecionar imagens
        frame_imagens = tk.Frame(frame, bg='#34495e')
        frame_imagens.pack(fill='x', pady=10)
        
        tk.Label(frame_imagens, text="🖼️ Imagens:", font=('Arial', 12), 
                bg='#34495e', fg='white').pack(side='left', padx=10, pady=10)
        
        self.btn_imagens = tk.Button(frame_imagens, text="Selecionar Imagens", 
                                  command=self.selecionar_imagens,
                                  bg='#e74c3c', fg='white', font=('Arial', 10, 'bold'),
                                  padx=20, pady=10)
        self.btn_imagens.pack(side='left', padx=10, pady=10)
        
        self.lbl_imagens = tk.Label(frame_imagens, text="Nenhuma imagem selecionada", 
                                  font=('Arial', 10), bg='#34495e', fg='#ecf0f1')
        self.lbl_imagens.pack(side='left', padx=10, pady=10)
        
        # Lista de produtos
        frame_lista = tk.Frame(frame, bg='#34495e')
        frame_lista.pack(fill='both', expand=True, pady=10)
        
        tk.Label(frame_lista, text="📋 Produtos encontrados:", font=('Arial', 12), 
                bg='#34495e', fg='white').pack(pady=10)
        
        # Treeview para produtos
        cols = ('Código', 'Descrição', 'Imagem Atual')
        self.tree = ttk.Treeview(frame_lista, columns=cols, show='headings', height=15)
        
        for col in cols:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=200)
        
        scrollbar = ttk.Scrollbar(frame_lista, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # Botões de ação
        frame_botoes = tk.Frame(frame, bg='#2c3e50')
        frame_botoes.pack(fill='x', pady=20)
        
        self.btn_mapear = tk.Button(frame_botoes, text="🔗 Mapear Imagens", 
                                 command=self.mapear_imagens,
                                 bg='#27ae60', fg='white', font=('Arial', 12, 'bold'),
                                 padx=30, pady=15)
        self.btn_mapear.pack(side='left', padx=10)
        
        self.btn_salvar = tk.Button(frame_botoes, text="💾 Salvar Tudo", 
                                command=self.salvar_tudo,
                                bg='#f39c12', fg='white', font=('Arial', 12, 'bold'),
                                padx=30, pady=15)
        self.btn_salvar.pack(side='left', padx=10)
        
    def selecionar_excel(self):
        arquivo = filedialog.askopenfilename(
            title="Selecione o arquivo Excel",
            filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
        )
        
        if arquivo:
            self.arquivo_excel = arquivo
            self.lbl_excel.config(text=os.path.basename(arquivo))
            self.carregar_produtos()
            
    def carregar_produtos(self):
        if not self.arquivo_excel:
            return
            
        try:
            wb = load_workbook(self.arquivo_excel, read_only=True)
            ws = wb.active
            
            # Limpar tree
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            # Carregar produtos
            for row_num in range(3, min(25, ws.max_row + 1)):
                row = next(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
                if len(row) >= 5 and row[4]:  # Tem código na coluna 4
                    codigo = str(row[4]).strip()
                    descricao = str(row[10]).strip()[:50] + "..." if len(row) > 10 and row[10] else "Sem descrição"
                    
                    self.tree.insert('', 'end', values=(codigo, descricao, ""))
            
            wb.close()
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar Excel: {e}")
            
    def selecionar_imagens(self):
        arquivos = filedialog.askopenfilenames(
            title="Selecione as imagens",
            filetypes=[("Imagens", "*.jpg *.jpeg *.png *.gif *.bmp")]
        )
        
        if arquivos:
            self.imagens_selecionadas = list(arquivos)
            self.lbl_imagens.config(text=f"{len(arquivos)} imagens selecionadas")
            
    def mapear_imagens(self):
        if not self.arquivo_excel or not self.imagens_selecionadas:
            messagebox.showwarning("Aviso", "Selecione o Excel e as imagens primeiro!")
            return
            
        # Janela de mapeamento
        janela_mapear = tk.Toplevel(self.janela)
        janela_mapear.title("🔗 Mapear Imagens")
        janela_mapear.geometry("600x400")
        janela_mapear.configure(bg='#2c3e50')
        
        tk.Label(janela_mapear, text="Associe cada imagem ao produto correspondente:",
                font=('Arial', 12), bg='#2c3e50', fg='white').pack(pady=20)
        
        frame_mapeamento = tk.Frame(janela_mapear, bg='#2c3e50')
        frame_mapeamento.pack(fill='both', expand=True, padx=20)
        
        self.mapeamentos = {}
        
        # Criar comboboxes para cada imagem
        for i, img_path in enumerate(self.imagens_selecionadas[:10]):  # Limitar a 10
            frame_img = tk.Frame(frame_mapeamento, bg='#34495e')
            frame_img.pack(fill='x', pady=5)
            
            tk.Label(frame_img, text=f"🖼️ {os.path.basename(img_path)}:",
                    font=('Arial', 10), bg='#34495e', fg='white').pack(side='left', padx=10)
            
            # Obter códigos dos produtos
            codigos = [self.tree.item(item)['values'][0] for item in self.tree.get_children()]
            
            combo = ttk.Combobox(frame_img, values=codigos, width=30)
            combo.pack(side='left', padx=10)
            
            self.mapeamentos[img_path] = combo
            
        tk.Button(janela_mapear, text="✅ Confirmar Mapeamento",
                 command=lambda: self.confirmar_mapeamento(janela_mapear),
                 bg='#27ae60', fg='white', font=('Arial', 12, 'bold'),
                 padx=20, pady=10).pack(pady=20)
        
    def confirmar_mapeamento(self, janela):
        # Processar mapeamentos
        for img_path, combo in self.mapeamentos.items():
            codigo_selecionado = combo.get()
            if codigo_selecionado:
                # Encontrar produto no tree e atualizar
                for item in self.tree.get_children():
                    if self.tree.item(item)['values'][0] == codigo_selecionado:
                        self.tree.item(item, values=(
                            codigo_selecionado,
                            self.tree.item(item)['values'][1],
                            os.path.basename(img_path)
                        ))
                        break
        
        janela.destroy()
        messagebox.showinfo("Sucesso", "Mapeamento realizado com sucesso!")
        
    def salvar_tudo(self):
        if not self.arquivo_excel:
            messagebox.showwarning("Aviso", "Selecione o arquivo Excel primeiro!")
            return
            
        # Obter nome do cliente
        cliente = os.path.basename(self.arquivo_excel).replace('.xlsx', '').replace('.xls', '')
        
        # Criar pasta do cliente
        pasta_cliente = os.path.join("imagens", cliente)
        if not os.path.exists(pasta_cliente):
            os.makedirs(pasta_cliente)
        
        imagens_salvas = 0
        
        # Salvar imagens mapeadas
        for item in self.tree.get_children():
            valores = self.tree.item(item)['values']
            codigo = valores[0]
            nome_imagem = valores[2]
            
            if nome_imagem:
                # Encontrar o arquivo original
                for img_path in self.imagens_selecionadas:
                    if os.path.basename(img_path) == nome_imagem:
                        # Copiar e renomear
                        novo_nome = f"{codigo}.jpg"
                        novo_caminho = os.path.join(pasta_cliente, novo_nome)
                        
                        # Converter para JPG se necessário
                        try:
                            img = Image.open(img_path)
                            if img.mode in ('RGBA', 'LA', 'P'):
                                img = img.convert('RGB')
                            img.save(novo_caminho, 'JPEG', quality=85)
                            imagens_salvas += 1
                        except Exception as e:
                            print(f"Erro ao salvar imagem {img_path}: {e}")
                        break
        
        messagebox.showinfo("Sucesso", f"Salvas {imagens_salvas} imagens na pasta {pasta_cliente}")
        
    def executar(self):
        self.janela.mainloop()

if __name__ == "__main__":
    app = UploadImagens()
    app.executar()
