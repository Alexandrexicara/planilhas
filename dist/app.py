from flask import Flask, render_template, request, jsonify, redirect, url_for, flash, send_file
import os
import sys
import sqlite3
import zipfile
import shutil
import subprocess
import time
import importlib
import threading
import webbrowser
import runpy
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import json
from datetime import datetime

# Importacao segura dos modulos desktop (podem falhar em ambiente server/headless)
MODULOS_IMPORTACAO = {}
ERROS_IMPORTACAO = {}


def importar_modulo(nome_modulo):
    try:
        modulo = importlib.import_module(nome_modulo)
        MODULOS_IMPORTACAO[nome_modulo] = True
        return modulo
    except Exception as e:
        MODULOS_IMPORTACAO[nome_modulo] = False
        ERROS_IMPORTACAO[nome_modulo] = str(e)
        print(f"[AVISO] Nao foi possivel importar {nome_modulo}: {e}")
        return None


sistema = importar_modulo('sistema')
sistema_plus = importar_modulo('sistema_plus')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
IS_VERCEL = bool(os.environ.get('VERCEL'))
IS_RENDER = bool(os.environ.get('RENDER'))
RUNTIME_DIR = os.path.join('/tmp', 'planilhas') if IS_VERCEL else BASE_DIR
UPLOAD_DIR = os.path.join(RUNTIME_DIR, 'uploads')
TEMP_IMAGES_DIR = os.path.join(RUNTIME_DIR, 'temp_images')
STATIC_UPLOADS_DIR = os.path.join(RUNTIME_DIR, 'static_uploads')
BUNDLED_DB_PATH = os.path.join(BASE_DIR, 'banco_plus.db')
DB_PATH = os.path.join(RUNTIME_DIR, 'banco_plus.db') if IS_VERCEL else BUNDLED_DB_PATH

app = Flask(__name__)
app.secret_key = 'sistema_plus_2024'
app.config['UPLOAD_FOLDER'] = UPLOAD_DIR
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Pastas de runtime (serverless so pode escrever em /tmp)
for runtime_path in [RUNTIME_DIR, UPLOAD_DIR, TEMP_IMAGES_DIR, STATIC_UPLOADS_DIR]:
    try:
        os.makedirs(runtime_path, exist_ok=True)
    except Exception as e:
        print(f"[AVISO] Nao foi possivel criar pasta {runtime_path}: {e}")

# Em Vercel, copia o banco empacotado para /tmp na primeira execucao
if IS_VERCEL and os.path.exists(BUNDLED_DB_PATH) and not os.path.exists(DB_PATH):
    try:
        shutil.copy2(BUNDLED_DB_PATH, DB_PATH)
    except Exception as e:
        print(f"[AVISO] Nao foi possivel copiar banco para runtime: {e}")

# Configuracao do sistema
SISTEMA_CONFIG = {
    'nome': 'planilhas.com',
    'valor_original': 5000.00,
    'valor_promocional': 4500.00,
    'desconto': 10,
    'versao': '2.0',
    'recursos': [
        'Importacao automatica de Excel',
        'Extracao automatica de imagens',
        'Catalogo web completo',
        'Gestao de produtos',
        'Relatorios detalhados'
    ]
}

BUILD_INFO = {
    'render': IS_RENDER,
    'service': os.environ.get('RENDER_SERVICE_NAME', ''),
    'vercel': IS_VERCEL,
    'region': os.environ.get('VERCEL_REGION', ''),
    'commit': os.environ.get('RENDER_GIT_COMMIT', '') or os.environ.get('VERCEL_GIT_COMMIT_SHA', ''),
}

DESKTOP_RELEASES_DIR = os.path.join(BASE_DIR, 'releases')
DESKTOP_POLICY = {
    'max_maquinas': 10,
    'atualizacao_requer_compra': True
}


def abrir_navegador_em_background(url, atraso_segundos=1.2):
    """Abre o navegador sem travar o startup do Flask."""
    def _abrir():
        time.sleep(atraso_segundos)
        try:
            webbrowser.open(url, new=1)
        except Exception as e:
            print(f"[AVISO] Nao foi possivel abrir navegador automaticamente: {e}")

    threading.Thread(target=_abrir, daemon=True).start()

def get_db_connection():
    """Conexao com o banco de dados"""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def get_latest_desktop_exe():
    """Retorna o executavel mais recente em `releases/`."""
    if not os.path.isdir(DESKTOP_RELEASES_DIR):
        return None

    exes = []
    for entry in os.scandir(DESKTOP_RELEASES_DIR):
        if entry.is_file() and entry.name.lower().endswith('.exe'):
            exes.append(entry.path)

    if not exes:
        return None

    latest_path = max(exes, key=os.path.getmtime)
    return {
        'path': latest_path,
        'filename': os.path.basename(latest_path),
        'updated_at': datetime.fromtimestamp(os.path.getmtime(latest_path)).strftime('%Y-%m-%d %H:%M:%S')
    }


def get_desktop_distribution_info():
    """Dados para exibir download .exe e limite de maquinas."""
    exe = get_latest_desktop_exe()
    return {
        'disponivel': bool(exe),
        'arquivo': exe['filename'] if exe else None,
        'atualizado_em': exe['updated_at'] if exe else None,
        'download_url': '/download-exe' if exe else None,
        'max_maquinas': DESKTOP_POLICY['max_maquinas'],
        'atualizacao_requer_compra': DESKTOP_POLICY['atualizacao_requer_compra']
    }

def extract_images_from_excel(excel_path, output_dir):
    """Extrai imagens do arquivo Excel automaticamente"""
    images = []
    
    try:
        with zipfile.ZipFile(excel_path, 'r') as z:
            for file in z.namelist():
                if file.startswith("xl/media/"):
                    name = os.path.basename(file)
                    path = os.path.join(output_dir, name)
                    
                    with open(path, "wb") as f:
                        f.write(z.read(file))
                    
                    images.append(path)
    except Exception as e:
        print(f"Erro ao extrair imagens: {e}")
    
    return images

def read_excel_with_images(excel_path):
    """Le Excel e associa imagens automaticamente"""
    temp_dir = TEMP_IMAGES_DIR
    os.makedirs(temp_dir, exist_ok=True)
    
    # 1. Extrair imagens
    images = extract_images_from_excel(excel_path, temp_dir)
    
    # 2. Ler dados do Excel
    workbook = load_workbook(excel_path)
    sheet = workbook.active
    
    rows = []
    headers = []
    
    # Pegar cabecalhos
    for cell in sheet[1]:
        if cell.value:
            headers.append(str(cell.value).strip())
    
    # Ler linhas de dados
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        row_data = {}
        for i, value in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = str(value) if value is not None else ""
        
        # 3. Associar imagem por ordem
        img_index = row_idx - 2  # Ajuste para indice zero
        if img_index < len(images):
            row_data['picture'] = images[img_index]
        else:
            row_data['picture'] = None
        
        rows.append(row_data)
    
    return rows, images

def save_image_to_static(image_path):
    """Move imagem para pasta static e retorna URL"""
    if not image_path or not os.path.exists(image_path):
        return None
    
    filename = os.path.basename(image_path)
    new_path = os.path.join(STATIC_UPLOADS_DIR, filename)
    
    shutil.copy(image_path, new_path)
    return new_path if IS_VERCEL else f"/static/uploads/{filename}"

def import_products_to_db(products_data):
    """Importa produtos para o banco de dados"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    imported_count = 0
    
    for product in products_data:
        try:
            # Salvar imagem se existir
            picture_url = None
            if product.get('picture'):
                picture_url = save_image_to_static(product['picture'])
            
            # Inserir no banco
            cursor.execute("""
                INSERT INTO produtos_plus (
                    cliente, arquivo_origem, codigo, descricao, peso, 
                    picture, data_importacao
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """, (
                product.get('cliente', 'Web'),
                'upload_web',
                product.get('codigo', ''),
                product.get('descricao', ''),
                product.get('peso', ''),
                picture_url,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ))
            
            imported_count += 1
            
        except Exception as e:
            print(f"Erro ao importar produto: {e}")
            continue
    
    conn.commit()
    conn.close()
    
    # Limpar pasta temporaria
    if os.path.exists(TEMP_IMAGES_DIR):
        shutil.rmtree(TEMP_IMAGES_DIR)
    
    return imported_count


def ambiente_desktop_disponivel():
    """Indica se o servidor consegue abrir janelas desktop (GUI)."""
    if os.name == 'nt':
        return True
    return bool(os.environ.get('DISPLAY') or os.environ.get('WAYLAND_DISPLAY'))


def nome_ambiente_execucao():
    """Nome amigavel do ambiente atual para mensagens."""
    if os.environ.get('RENDER'):
        return 'Render'
    if os.environ.get('VERCEL'):
        return 'Vercel'
    if os.name == 'nt':
        return 'Windows'
    return 'Servidor Linux sem interface grafica'


def abrir_janela_sistema(script_name, nome_exibicao):
    """Abre um sistema desktop em uma nova janela sem bloquear o Flask."""
    script_path = os.path.join(BASE_DIR, script_name)
    module_name = os.path.splitext(script_name)[0]
    running_frozen = bool(getattr(sys, 'frozen', False))

    if not os.path.exists(script_path) and not running_frozen:
        raise FileNotFoundError(f"Arquivo nao encontrado: {script_name}")

    if not ambiente_desktop_disponivel():
        ambiente = nome_ambiente_execucao()
        raise RuntimeError(
            f"Nao e possivel abrir janela desktop no ambiente {ambiente}. Use a interface web."
        )

    popen_kwargs = {
        'cwd': BASE_DIR,
    }

    if os.name == 'nt':
        pythonw_path = os.path.join(os.path.dirname(sys.executable), 'pythonw.exe')
        executable = pythonw_path if os.path.exists(pythonw_path) else sys.executable
        creationflags = getattr(subprocess, 'CREATE_NEW_CONSOLE', 0)
        if creationflags:
            popen_kwargs['creationflags'] = creationflags
    else:
        executable = sys.executable
        popen_kwargs['start_new_session'] = True

    if os.path.exists(script_path):
        cmd = [executable, script_path]
    elif running_frozen:
        cmd = [executable, "--run-module", module_name]
    else:
        raise FileNotFoundError(f"Arquivo nao encontrado: {script_name}")

    processo = subprocess.Popen(cmd, **popen_kwargs)
    time.sleep(0.6)
    if processo.poll() is not None:
        raise RuntimeError(
            f"{nome_exibicao} encerrou logo apos iniciar (codigo {processo.returncode})."
        )

    return {
        'nome': nome_exibicao,
        'arquivo': script_name,
        'pid': processo.pid,
        'status': 'sucesso',
        'mensagem': f'{nome_exibicao} aberto em nova janela.'
    }

@app.route('/')
def index():
    """Pagina inicial atraente"""
    return render_template('index.html', config=SISTEMA_CONFIG)


@app.route('/download-exe')
def download_exe():
    """Baixa o executavel desktop mais recente publicado em releases/."""
    exe = get_latest_desktop_exe()
    if not exe:
        return (
            "Arquivo .exe ainda nao foi publicado. "
            "Adicione o instalador em releases/ para liberar o download.",
            404
        )

    return send_file(exe['path'], as_attachment=True, download_name=exe['filename'])

@app.route('/executar-sistema-legado')
def executar_sistema_legado():
    """Executa funcoes dos modulos dentro do mesmo processo Flask."""
    try:
        dados_sistema = {
            'status': 'sucesso',
            'mensagem': 'Sistema iniciado com sucesso!',
            'timestamp': datetime.now().isoformat(),
            'modulos_carregados': [],
            'distribuicao_desktop': get_desktop_distribution_info()
        }

        # Modulo sistema
        try:
            if sistema is not None:
                dados_sistema['modulos_carregados'].append('[OK] sistema.py carregado')
            else:
                erro = ERROS_IMPORTACAO.get('sistema', 'modulo nao carregado')
                dados_sistema['modulos_carregados'].append(f'[ERRO] sistema.py: {erro[:80]}')
        except Exception as e:
            dados_sistema['modulos_carregados'].append(f'[ERRO] sistema.py: {str(e)[:50]}')

        # Modulo sistema_plus
        try:
            if sistema_plus is not None and hasattr(sistema_plus, 'contar_produtos_plus'):
                dados_sistema['total_produtos'] = sistema_plus.contar_produtos_plus()
            if sistema_plus is not None and hasattr(sistema_plus, 'contar_planilhas_plus'):
                dados_sistema['total_planilhas'] = sistema_plus.contar_planilhas_plus()

            if sistema_plus is not None:
                dados_sistema['modulos_carregados'].append('[OK] sistema_plus.py carregado')
            else:
                erro = ERROS_IMPORTACAO.get('sistema_plus', 'modulo nao carregado')
                dados_sistema['modulos_carregados'].append(f'[ERRO] sistema_plus.py: {erro[:80]}')
        except Exception as e:
            dados_sistema['modulos_carregados'].append(f'[ERRO] sistema_plus.py: {str(e)[:50]}')

        return render_template(
            'sistema.html',
            config=SISTEMA_CONFIG,
            resultado=dados_sistema
        )
    except Exception as e:
        return render_template('iniciando.html', erro=str(e))


@app.route('/executar-sistema')
def executar_sistema():
    """Abre as janelas do sistema original e/ou PLUS a partir do Flask."""
    try:
        alvo = request.args.get('target', 'both').lower()
        scripts_por_alvo = {
            'original': [('sistema.py', 'Sistema Original')],
            'plus': [('sistema_plus.py', 'Sistema Plus')],
            'menu': [('menu_principal.py', 'Menu Principal')],
            'both': [
                ('sistema.py', 'Sistema Original'),
                ('sistema_plus.py', 'Sistema Plus')
            ],
            'all': [
                ('menu_principal.py', 'Menu Principal'),
                ('sistema.py', 'Sistema Original'),
                ('sistema_plus.py', 'Sistema Plus')
            ]
        }

        scripts_para_abrir = scripts_por_alvo.get(alvo, scripts_por_alvo['both'])
        ambiente_gui = ambiente_desktop_disponivel()
        ambiente_nome = nome_ambiente_execucao()
        dados_sistema = {
            'status': 'sucesso',
            'mensagem': 'As janelas do sistema foram abertas com sucesso.',
            'timestamp': datetime.now().isoformat(),
            'modulos_carregados': [],
            'janelas_abertas': [],
            'alvo': alvo,
            'ambiente_gui': ambiente_gui,
            'ambiente_nome': ambiente_nome,
            'build_info': BUILD_INFO,
            'distribuicao_desktop': get_desktop_distribution_info()
        }

        for script_name, nome_exibicao in scripts_para_abrir:
            try:
                janela = abrir_janela_sistema(script_name, nome_exibicao)
                dados_sistema['janelas_abertas'].append(janela)
                dados_sistema['modulos_carregados'].append(f'[OK] {script_name} aberto em nova janela')
            except Exception as e:
                dados_sistema['status'] = 'parcial'
                dados_sistema['janelas_abertas'].append({
                    'nome': nome_exibicao,
                    'arquivo': script_name,
                    'status': 'erro',
                    'mensagem': str(e)
                })
                dados_sistema['modulos_carregados'].append(f'[ERRO] {script_name}: {str(e)[:80]}')

        try:
            if sistema_plus is not None and hasattr(sistema_plus, 'contar_produtos_plus'):
                dados_sistema['total_produtos'] = sistema_plus.contar_produtos_plus()
            if sistema_plus is not None and hasattr(sistema_plus, 'contar_planilhas_plus'):
                dados_sistema['total_planilhas'] = sistema_plus.contar_planilhas_plus()

            if sistema_plus is not None:
                dados_sistema['modulos_carregados'].append('[OK] Estatisticas do sistema_plus.py carregadas')
            else:
                erro = ERROS_IMPORTACAO.get('sistema_plus', 'modulo nao carregado')
                dados_sistema['modulos_carregados'].append(
                    f'[ERRO] Estatisticas indisponiveis (sistema_plus.py): {erro[:80]}'
                )
        except Exception as e:
            dados_sistema['modulos_carregados'].append(f'[ERRO] Estatisticas do sistema_plus.py: {str(e)[:80]}')

        if not ambiente_gui:
            dados_sistema['status'] = 'erro'
            dados_sistema['mensagem'] = (
                f"Ambiente {ambiente_nome} nao suporta janelas desktop. "
                "Use apenas as rotas web."
            )

        if dados_sistema['janelas_abertas'] and all(
            item.get('status') == 'erro' for item in dados_sistema['janelas_abertas']
        ):
            dados_sistema['status'] = 'erro'
            if ambiente_gui:
                dados_sistema['mensagem'] = 'Nao foi possivel abrir as janelas solicitadas.'
        elif ambiente_gui and any(item.get('status') == 'erro' for item in dados_sistema['janelas_abertas']):
            dados_sistema['mensagem'] = 'Algumas janelas abriram e outras falharam.'

        return render_template('sistema.html',
                             config=SISTEMA_CONFIG,
                             resultado=dados_sistema)
    except Exception as e:
        return render_template('iniciando.html', erro=str(e))



@app.route('/upload')
def upload_page():
    """Pagina de upload"""
    return render_template('upload.html', config=SISTEMA_CONFIG)

@app.route('/api/upload', methods=['POST'])
def upload_excel():
    """API para upload e processamento de Excel"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'Nenhum arquivo enviado'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Nome de arquivo invalido'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Arquivo deve ser Excel (.xlsx ou .xls)'}), 400
        
        # Salvar arquivo temporario
        filename = secure_filename(file.filename)
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(temp_path)
        
        # Processar Excel
        products_data, images = read_excel_with_images(temp_path)
        
        # Importar para banco
        imported_count = import_products_to_db(products_data)
        
        # Limpar arquivo temporario
        os.remove(temp_path)
        
        return jsonify({
            'success': True,
            'message': f'Importacao concluida com sucesso!',
            'imported_count': imported_count,
            'images_found': len(images),
            'products_count': len(products_data)
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/catalog')
def catalog():
    """Catalogo de produtos"""
    conn = get_db_connection()
    products = conn.execute("""
        SELECT * FROM produtos_plus 
        ORDER BY data_importacao DESC 
        LIMIT 100
    """).fetchall()
    conn.close()
    
    return render_template('catalog.html', products=products, config=SISTEMA_CONFIG)

@app.route('/api/stats')
def get_stats():
    """API de estatisticas"""
    conn = get_db_connection()
    
    stats = {
        'total_products': conn.execute("SELECT COUNT(*) FROM produtos_plus").fetchone()[0],
        'total_clients': conn.execute("SELECT COUNT(DISTINCT cliente) FROM produtos_plus").fetchone()[0],
        'recent_imports': conn.execute("""
            SELECT COUNT(*) FROM produtos_plus 
            WHERE data_importacao >= date('now', '-7 days')
        """).fetchone()[0]
    }
    
    conn.close()
    return jsonify(stats)


def executar_modulo_desktop(module_name):
    """Executa um modulo desktop como se fosse script (__main__)."""
    try:
        runpy.run_module(module_name, run_name="__main__")
        return 0
    except Exception as e:
        print(f"[ERRO] Falha ao executar modulo {module_name}: {e}")
        return 1


def executar_app():
    """Inicializa o Flask em modo desktop/web sem abrir console de debug."""
    port = int(os.environ.get("PORT", 5000))
    debug_flag = str(os.environ.get("FLASK_DEBUG", "0")).lower() in {"1", "true", "yes", "on"}
    host = '0.0.0.0' if (IS_VERCEL or IS_RENDER) else '127.0.0.1'
    auto_open_browser = (
        str(os.environ.get("AUTO_OPEN_BROWSER", "1")).lower() in {"1", "true", "yes", "on"}
    )

    if auto_open_browser and host == '127.0.0.1':
        abrir_navegador_em_background(f"http://127.0.0.1:{port}")

    app.run(
        debug=debug_flag,
        host=host,
        port=port,
        use_reloader=False
    )


if __name__ == '__main__':
    if len(sys.argv) >= 3 and sys.argv[1] == "--run-module":
        sys.exit(executar_modulo_desktop(sys.argv[2]))
    executar_app()

